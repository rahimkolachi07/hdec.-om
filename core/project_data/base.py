"""
project_data/base.py — Shared directory helpers for per-project data storage.

Disk layout:
  projects_data/
    <country_id>/
      <project_id>/
        manpower/
          data.json
        cmms/
          activities.json
          permits/
            <permit_id>.json
          handover/
            <handover_id>.json
"""
import json, io, openpyxl
from pathlib import Path
from datetime import datetime

BASE = Path(__file__).resolve().parent.parent.parent / 'projects_data'


def get_project_dir(cid: str, pid: str) -> Path:
    d = BASE / cid / pid
    d.mkdir(parents=True, exist_ok=True)
    return d


def get_module_dir(cid: str, pid: str, *parts: str) -> Path:
    d = get_project_dir(cid, pid)
    for p in parts:
        d = d / p
    d.mkdir(parents=True, exist_ok=True)
    return d


def load_json(path: Path):
    """Load JSON from a path. Returns None if missing or invalid."""
    if not path.exists():
        return None
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None


def save_json(path: Path, data):
    """Write JSON to a path, creating parent dirs as needed."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def excel_style_header(ws, headers: list, col_widths: dict = None):
    """Write a styled header row to an openpyxl worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1a3a5c')
        cell.alignment = Alignment(horizontal='center')
    if col_widths:
        for letter, w in col_widths.items():
            ws.column_dimensions[letter].width = w


# ══════════════════════════════════════════════════════════════════════════════
# BLANK TEMPLATE DOWNLOADS
# ══════════════════════════════════════════════════════════════════════════════

def get_blank_template(module: str) -> bytes:
    """Return a blank Excel import template for the given module."""
    wb = openpyxl.Workbook()
    from openpyxl.styles import Font

    def make_sheet(ws, headers, sample_row=None, widths=None):
        excel_style_header(ws, headers, widths)
        if sample_row:
            ws.append(sample_row)

    if module == 'manpower':
        wb.remove(wb.active)
        ws_eng = wb.create_sheet('Engineers')
        make_sheet(
            ws_eng,
            ['Name', 'Role', 'Department', '2025-01-01', '2025-01-02', '...more dates'],
            sample_row=['Ahmed Ali', 'Operation Engineer', 'Operations', 'Day', 'Night', ''],
            widths={'A': 24, 'B': 22, 'C': 20, 'D': 12, 'E': 12},
        )
        ws_tech = wb.create_sheet('Technicians')
        make_sheet(
            ws_tech,
            ['Name', 'Role', 'Department', '2025-01-01', '2025-01-02', '...more dates'],
            sample_row=['Khalid', 'Technician', '', 'Day', 'OFF', ''],
            widths={'A': 24, 'B': 22, 'C': 20},
        )
        ws_inst = wb.create_sheet('Instructions')
        ws_inst.append(['MANPOWER IMPORT INSTRUCTIONS'])
        ws_inst.append([])
        ws_inst.append(['Sheet "Engineers": Enter engineering staff schedules'])
        ws_inst.append(['Sheet "Technicians": Enter technician schedules'])
        ws_inst.append(['Columns: Name, Role, Department, then one column per date (YYYY-MM-DD)'])
        ws_inst.append(['Shift values: Day | Night | General | OFF | Leave | Rest'])
        ws_inst['A1'].font = Font(bold=True, size=14)

    elif module == 'activities':
        ws = wb.active
        ws.title = 'Activities'
        make_sheet(
            ws,
            ['Name', 'Equipment', 'Location', 'Type', 'Frequency', 'Description'],
            sample_row=['PV Module Inspection', 'PV Strings', 'Zone A', 'PM', 'Monthly', 'Visual inspection'],
            widths={'A': 28, 'B': 22, 'C': 20, 'D': 12, 'E': 14, 'F': 40},
        )
        ws_inst = wb.create_sheet('Instructions')
        ws_inst.append(['ACTIVITIES IMPORT INSTRUCTIONS'])
        ws_inst.append(['Type: PM (Preventive) | CM (Corrective) | Inspection'])
        ws_inst.append(['Frequency: Daily | Weekly | Monthly | Quarterly | Annually'])
        ws_inst['A1'].font = Font(bold=True, size=14)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
