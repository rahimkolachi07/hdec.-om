"""
CMMS utilities for monthly activities, web checklist editing, evidence photos,
and ZIP generation.

Primary data lives in ``cmms_data/``. Checklist workbooks are discovered from:
1. ``cmms_data/checklists/``
2. ``cmms_data/``
3. project-root ``Checklists/``
4. legacy ``media/cmms/checklists/``
"""
from __future__ import annotations

import io
import json
import re
import shutil
import uuid
import zipfile
import csv
from datetime import date, datetime, time, timedelta
from pathlib import Path
from urllib.parse import parse_qs, urlencode, urlparse
from urllib.request import Request, urlopen

from django.conf import settings

BASE_DIR = Path(
    getattr(settings, 'BASE_DIR', Path(__file__).resolve().parent.parent)
).resolve()
CMMS_DATA_DIR = Path(
    getattr(settings, 'CMMS_DATA_DIR', Path(__file__).resolve().parent.parent / 'cmms_data')
).resolve()
MEDIA_ROOT = Path(
    getattr(settings, 'MEDIA_ROOT', Path(__file__).resolve().parent.parent / 'media')
).resolve()
MEDIA_CHECKLISTS_DIR = MEDIA_ROOT / 'cmms' / 'checklists'
LOCAL_CHECKLISTS_DIR = CMMS_DATA_DIR / 'checklists'
ROOT_CHECKLISTS_DIR = BASE_DIR / 'Checklists'
PHOTOS_DIR = MEDIA_ROOT / 'cmms' / 'photos'
CMMS_CM_REPORT_TEMPLATE_URL = str(
    getattr(
        settings,
        'CMMS_CM_REPORT_TEMPLATE_URL',
        'https://docs.google.com/spreadsheets/d/1Cvy43EKlDkVvBrFfa0f1b5eMZv04C273GGEOU5Oo234/edit?usp=sharing',
    ) or ''
).strip()
CHECKLIST_ACTIVITY_CACHE_SECONDS = max(
    30,
    int(getattr(settings, 'CMMS_CHECKLIST_CACHE_SECONDS', 300) or 300),
)

_CHECKLIST_ACTIVITY_CACHE: dict[str, object] = {
    'expires_at': None,
    'activities': None,
}

for directory in (CMMS_DATA_DIR, LOCAL_CHECKLISTS_DIR, ROOT_CHECKLISTS_DIR, MEDIA_CHECKLISTS_DIR, PHOTOS_DIR):
    directory.mkdir(parents=True, exist_ok=True)

ACTIVITIES_FILE = CMMS_DATA_DIR / 'activities.json'
RECORDS_FILE = CMMS_DATA_DIR / 'records.json'
CHECKLIST_EXTENSIONS = {'.xlsx', '.xlsm', '.xls'}
ONE_TIME_FREQUENCIES = {'', 'one-time', 'one time', 'once', 'single'}
MONTH_STEP_FREQUENCIES = {
    'monthly': 1,
    'quarterly': 3,
    'half-yearly': 6,
    'half yearly': 6,
    'yearly': 12,
}


def _load(path: Path) -> list:
    if not path.exists():
        return []
    try:
        with path.open('r', encoding='utf-8') as handle:
            data = json.load(handle)
            return data if isinstance(data, list) else []
    except Exception:
        return []


def _save(path: Path, data: list) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8') as handle:
        json.dump(data, handle, indent=2, ensure_ascii=False, default=str)


def _record_sort_key(record: dict) -> tuple:
    return (
        bool(record.get('completed')),
        record.get('completed_at') or '',
        record.get('started_at') or '',
        record.get('created_at') or '',
        record.get('id') or '',
    )


def _serialize_checklist_ref(path: Path) -> str:
    path = Path(path).resolve()
    if path.is_relative_to(CMMS_DATA_DIR):
        return f"cmms_data:{path.relative_to(CMMS_DATA_DIR).as_posix()}"
    if path.is_relative_to(BASE_DIR):
        return f"project:{path.relative_to(BASE_DIR).as_posix()}"
    if path.is_relative_to(MEDIA_ROOT):
        return f"media:{path.relative_to(MEDIA_ROOT).as_posix()}"
    return str(path)


def resolve_checklist_path(checklist_ref: str | None) -> Path | None:
    if not checklist_ref:
        return None

    ref = str(checklist_ref).strip()
    if not ref:
        return None

    candidate = Path(ref)
    if candidate.is_absolute():
        return candidate if candidate.exists() else None

    if ':' in ref:
        prefix, rel = ref.split(':', 1)
        rel_path = Path(rel)
        if prefix == 'cmms_data':
            path = CMMS_DATA_DIR / rel_path
            return path if path.exists() else None
        if prefix == 'project':
            path = BASE_DIR / rel_path
            return path if path.exists() else None
        if prefix == 'media':
            path = MEDIA_ROOT / rel_path
            return path if path.exists() else None

    direct_candidates = [
        CMMS_DATA_DIR / ref,
        LOCAL_CHECKLISTS_DIR / ref,
        ROOT_CHECKLISTS_DIR / ref,
        MEDIA_ROOT / ref,
        MEDIA_CHECKLISTS_DIR / Path(ref).name,
        CMMS_DATA_DIR / Path(ref).name,
        LOCAL_CHECKLISTS_DIR / Path(ref).name,
        ROOT_CHECKLISTS_DIR / Path(ref).name,
    ]
    for path in direct_candidates:
        if path.exists():
            return path
    return None


def _iter_checklist_paths() -> list[Path]:
    seen: set[Path] = set()
    files: list[Path] = []
    for root in (LOCAL_CHECKLISTS_DIR, CMMS_DATA_DIR, ROOT_CHECKLISTS_DIR, MEDIA_CHECKLISTS_DIR):
        if not root.exists():
            continue
        for path in sorted(root.rglob('*')):
            if not path.is_file() or path.suffix.lower() not in CHECKLIST_EXTENSIONS:
                continue
            resolved = path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            files.append(resolved)
    return files


def save_checklist_file(uploaded_file) -> dict:
    ext = Path(uploaded_file.name).suffix.lower()
    if ext not in CHECKLIST_EXTENSIONS:
        raise ValueError('Checklist must be an Excel file (.xlsx, .xlsm, .xls).')

    stem = re.sub(r'[^A-Za-z0-9._ -]+', '', Path(uploaded_file.name).stem).strip(' .') or 'checklist'
    filename = f'{stem}_{datetime.now().strftime("%Y%m%d%H%M%S")}{ext}'
    dest = LOCAL_CHECKLISTS_DIR / filename

    with dest.open('wb') as handle:
        for chunk in uploaded_file.chunks():
            handle.write(chunk)

    return {
        'name': dest.name,
        'rel_path': _serialize_checklist_ref(dest),
    }


def _tokenize(text: str) -> set[str]:
    return {
        token
        for token in re.findall(r'[a-z0-9]+', (text or '').lower())
        if len(token) > 1
    }


def suggest_checklist_path(activity: dict | None) -> Path | None:
    if not activity:
        return None

    name = str(activity.get('name', '') or '').strip().lower()
    frequency = str(activity.get('frequency', '') or '').strip().lower()
    equipment = str(activity.get('equipment', '') or '').strip().lower()
    query = ' '.join(part for part in (name, frequency, equipment) if part)
    query_tokens = _tokenize(query)
    if not query_tokens:
        return None

    best_path = None
    best_score = 0
    for path in _iter_checklist_paths():
        stem = path.stem.lower()
        stem_tokens = _tokenize(stem)
        overlap = query_tokens & stem_tokens
        score = len(overlap) * 10
        if name and name == stem:
            score += 100
        elif name and name in stem:
            score += 40
        elif query_tokens and query_tokens.issubset(stem_tokens):
            score += 30
        if frequency and frequency in stem_tokens:
            score += 5
        if equipment and equipment in stem_tokens:
            score += 3
        if score > best_score:
            best_score = score
            best_path = path

    return best_path if best_score >= 15 else None


def ensure_activity_checklist(activity: dict | None) -> dict | None:
    if not activity:
        return None

    original_scheduled_date = activity.get('scheduled_date', '')
    original_month = activity.get('month', '')
    original_series_start = activity.get('series_start_date', '')
    original_recurring = activity.get('is_recurring')

    checklist_path = resolve_checklist_path(activity.get('checklist_file', ''))
    if checklist_path and activity.get('checklist_file') != _serialize_checklist_ref(checklist_path):
        updated = update_activity(activity['id'], {'checklist_file': _serialize_checklist_ref(checklist_path)}) or activity
        if updated is not activity:
            updated = dict(updated)
            updated['scheduled_date'] = original_scheduled_date or updated.get('scheduled_date', '')
            updated['month'] = original_month or updated.get('month', '')
            if original_series_start:
                updated['series_start_date'] = original_series_start
            if original_recurring is not None:
                updated['is_recurring'] = original_recurring
        return updated
    if checklist_path:
        return activity

    suggested = suggest_checklist_path(activity)
    if not suggested:
        return activity
    updated = update_activity(activity['id'], {'checklist_file': _serialize_checklist_ref(suggested)}) or activity
    if updated is not activity:
        updated = dict(updated)
        updated['scheduled_date'] = original_scheduled_date or updated.get('scheduled_date', '')
        updated['month'] = original_month or updated.get('month', '')
        if original_series_start:
            updated['series_start_date'] = original_series_start
        if original_recurring is not None:
            updated['is_recurring'] = original_recurring
    return updated


def get_activities() -> list:
    return _load(ACTIVITIES_FILE)


def get_activity(activity_id: str) -> dict | None:
    return next((activity for activity in get_activities() if activity.get('id') == activity_id), None)


def _activity_type(value: str) -> str:
    return 'CM' if str(value or '').strip().upper() == 'CM' else 'PM'


def _normalize_frequency(value: str) -> str:
    normalized = str(value or '').strip().lower().replace('_', '-')
    return re.sub(r'\s+', '-', normalized)


def _parse_iso_date(value: str) -> date | None:
    raw = str(value or '').strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw[:10])
    except Exception:
        return None


def _days_in_month(year: int, month: int) -> int:
    if month == 12:
        return (date(year + 1, 1, 1) - date(year, month, 1)).days
    return (date(year, month + 1, 1) - date(year, month, 1)).days


def _add_months(source: date, months: int) -> date:
    absolute_month = (source.year * 12) + source.month - 1 + months
    year, month_index = divmod(absolute_month, 12)
    month = month_index + 1
    day = min(source.day, _days_in_month(year, month))
    return date(year, month, day)


def _month_bounds(month: str) -> tuple[date, date] | tuple[None, None]:
    try:
        month_start = datetime.strptime(str(month or '').strip(), '%Y-%m').date()
        month_start = month_start.replace(day=1)
    except Exception:
        return None, None
    month_end = _add_months(month_start, 1) - timedelta(days=1)
    return month_start, month_end


def _activity_occurrence(activity: dict, occurrence_date: date) -> dict:
    base = dict(activity)
    base_date = str(activity.get('scheduled_date', '') or '')
    base['series_start_date'] = base_date
    base['scheduled_date'] = occurrence_date.isoformat()
    base['month'] = occurrence_date.strftime('%Y-%m')
    base['is_recurring'] = occurrence_date.isoformat() != base_date
    return base


def _excluded_dates(activity: dict | None) -> set[str]:
    return {
        str(value or '').strip()
        for value in (activity or {}).get('excluded_dates', []) or []
        if str(value or '').strip()
    }


def activity_occurs_on_date(activity: dict, scheduled_date: str) -> bool:
    target_date = _parse_iso_date(scheduled_date)
    start_date = _parse_iso_date(activity.get('scheduled_date', ''))
    if not target_date or not start_date or target_date < start_date:
        return False
    if target_date.isoformat() in _excluded_dates(activity):
        return False

    if _activity_type(activity.get('type', 'PM')) != 'PM':
        return target_date == start_date

    frequency = _normalize_frequency(activity.get('frequency', ''))
    if frequency in ONE_TIME_FREQUENCIES:
        return target_date == start_date
    if frequency == 'daily':
        return True
    if frequency == 'weekly':
        return (target_date - start_date).days % 7 == 0
    step_months = MONTH_STEP_FREQUENCIES.get(frequency)
    if step_months:
        months_between = (target_date.year - start_date.year) * 12 + (target_date.month - start_date.month)
        if months_between < 0 or months_between % step_months != 0:
            return False
        return _add_months(start_date, months_between) == target_date
    return target_date == start_date


def _expand_activity_for_month(activity: dict, month: str) -> list:
    month_start, month_end = _month_bounds(month)
    start_date = _parse_iso_date(activity.get('scheduled_date', ''))
    if not month_start or not month_end or not start_date or month_end < start_date:
        return []
    excluded_dates = _excluded_dates(activity)

    if _activity_type(activity.get('type', 'PM')) != 'PM':
        if month_start <= start_date <= month_end and start_date.isoformat() not in excluded_dates:
            return [_activity_occurrence(activity, start_date)]
        return []

    frequency = _normalize_frequency(activity.get('frequency', ''))
    occurrences: list[dict] = []

    if frequency in ONE_TIME_FREQUENCIES:
        if month_start <= start_date <= month_end and start_date.isoformat() not in excluded_dates:
            occurrences.append(_activity_occurrence(activity, start_date))
        return occurrences

    if frequency == 'daily':
        current = max(start_date, month_start)
        while current <= month_end:
            if current.isoformat() not in excluded_dates:
                occurrences.append(_activity_occurrence(activity, current))
            current += timedelta(days=1)
        return occurrences

    if frequency == 'weekly':
        current = start_date
        if current < month_start:
            delta_days = (month_start - current).days
            current += timedelta(days=(delta_days // 7) * 7)
            while current < month_start:
                current += timedelta(days=7)
        while current <= month_end:
            if current.isoformat() not in excluded_dates:
                occurrences.append(_activity_occurrence(activity, current))
            current += timedelta(days=7)
        return occurrences

    step_months = MONTH_STEP_FREQUENCIES.get(frequency)
    if step_months:
        occurrence_index = 0
        months_between = (month_start.year - start_date.year) * 12 + (month_start.month - start_date.month)
        if months_between > 0:
            occurrence_index = months_between // step_months

        current = _add_months(start_date, occurrence_index * step_months)
        while current < month_start:
            occurrence_index += 1
            current = _add_months(start_date, occurrence_index * step_months)
        while current <= month_end:
            if current.isoformat() not in excluded_dates:
                occurrences.append(_activity_occurrence(activity, current))
            occurrence_index += 1
            current = _add_months(start_date, occurrence_index * step_months)
        return occurrences

    if month_start <= start_date <= month_end and start_date.isoformat() not in excluded_dates:
        occurrences.append(_activity_occurrence(activity, start_date))
    return occurrences


def get_activities_for_month(month: str) -> list:
    activities: list[dict] = []
    for activity in get_activities():
        base_activity = ensure_activity_checklist(activity) or activity
        activities.extend(_expand_activity_for_month(base_activity, month))
    return sorted(activities, key=lambda activity: (activity.get('scheduled_date', ''), activity.get('name', '')))


def get_activities_for_date(date: str) -> list:
    activities: list[dict] = []
    target_date = _parse_iso_date(date)
    if not target_date:
        return activities
    for activity in get_activities():
        base_activity = ensure_activity_checklist(activity) or activity
        if activity_occurs_on_date(base_activity, target_date.isoformat()):
            activities.append(_activity_occurrence(base_activity, target_date))
    return sorted(activities, key=lambda activity: activity.get('name', ''))


def create_activity(data: dict) -> dict:
    activities = get_activities()
    checklist_ref = data.get('checklist_file', '')
    resolved = resolve_checklist_path(checklist_ref)
    if resolved:
        checklist_ref = _serialize_checklist_ref(resolved)

    act_type = data.get('type', 'PM').upper()
    if act_type not in ('PM', 'CM'):
        act_type = 'PM'
    frequency = data.get('frequency', '')
    if act_type == 'CM':
        frequency = 'one-time'
    activity = {
        'id': str(uuid.uuid4()),
        'type': act_type,
        'month': data.get('month', ''),
        'scheduled_date': data.get('scheduled_date', ''),
        'name': data.get('name', ''),
        'equipment': data.get('equipment', ''),
        'location': data.get('location', ''),
        'frequency': frequency,
        'checklist_file': checklist_ref,
        'assigned_engineer': data.get('assigned_engineer', ''),
        'assigned_technician': data.get('assigned_technician', ''),
        'notes': data.get('notes', ''),
        'excluded_dates': [],
        'created_at': datetime.now().isoformat(),
        'created_by': data.get('created_by', ''),
    }
    if not activity['checklist_file']:
        suggested = suggest_checklist_path(activity)
        if suggested:
            activity['checklist_file'] = _serialize_checklist_ref(suggested)
    activities.append(activity)
    _save(ACTIVITIES_FILE, activities)
    return activity


def update_activity(activity_id: str, data: dict) -> dict | None:
    activities = get_activities()
    for activity in activities:
        if activity.get('id') != activity_id:
            continue
        updates = {key: value for key, value in data.items() if key != 'id'}
        scheduled_date = str(updates.get('scheduled_date', '') or '').strip()
        if scheduled_date:
            updates['month'] = scheduled_date[:7]
        if 'excluded_dates' in updates:
            updates['excluded_dates'] = sorted({
                str(value or '').strip()
                for value in (updates.get('excluded_dates') or [])
                if str(value or '').strip()
            })
        if 'checklist_file' in updates:
            resolved = resolve_checklist_path(updates.get('checklist_file'))
            updates['checklist_file'] = _serialize_checklist_ref(resolved) if resolved else updates.get('checklist_file', '')
        activity.update(updates)
        _save(ACTIVITIES_FILE, activities)
        return activity
    return None


def _delete_activity_occurrence_data(activity_id: str, scheduled_date: str) -> None:
    clean_date = str(scheduled_date or '').strip()
    if not clean_date:
        return

    records = get_records()
    deleted_records = [
        record for record in records
        if record.get('activity_id') == activity_id and record.get('date') == clean_date
    ]
    if deleted_records:
        _save(
            RECORDS_FILE,
            [
                record for record in records
                if not (record.get('activity_id') == activity_id and record.get('date') == clean_date)
            ],
        )
        for record in deleted_records:
            record_dir = PHOTOS_DIR / str(record.get('id', '')).strip()
            if record_dir.exists():
                shutil.rmtree(record_dir, ignore_errors=True)

    record_ids = {
        str(record.get('id', '')).strip()
        for record in deleted_records
        if str(record.get('id', '')).strip()
    }
    permits_file = CMMS_DATA_DIR / 'permits.json'
    permits = _load(permits_file)
    if permits:
        remaining_permits = [
            permit for permit in permits
            if not (
                permit.get('activity_id') == activity_id
                and str(permit.get('scheduled_date', '') or '').strip() == clean_date
            )
            and str(permit.get('record_id', '')).strip() not in record_ids
        ]
        if len(remaining_permits) != len(permits):
            _save(permits_file, remaining_permits)


def delete_activity_occurrence(activity_id: str, scheduled_date: str) -> bool:
    activity = get_activity(activity_id)
    clean_date = str(scheduled_date or '').strip()
    if not activity or not clean_date:
        return False

    frequency = _normalize_frequency(activity.get('frequency', ''))
    is_recurring_pm = (
        _activity_type(activity.get('type', 'PM')) == 'PM'
        and frequency not in ONE_TIME_FREQUENCIES
    )

    if not is_recurring_pm:
        return delete_activity(activity_id)

    excluded_dates = _excluded_dates(activity)
    if clean_date not in excluded_dates:
        excluded_dates.add(clean_date)
        update_activity(activity_id, {'excluded_dates': sorted(excluded_dates)})
    _delete_activity_occurrence_data(activity_id, clean_date)
    return True


def delete_activity(activity_id: str) -> bool:
    activities = get_activities()
    target = next((activity for activity in activities if activity.get('id') == activity_id), None)
    if not target:
        return False
    remaining = [activity for activity in activities if activity.get('id') != activity_id]
    _save(ACTIVITIES_FILE, remaining)

    records = get_records()
    scheduled_dates = {
        str(record.get('date', '')).strip()
        for record in records
        if record.get('activity_id') == activity_id and str(record.get('date', '')).strip()
    }
    scheduled_dates.update(_excluded_dates(target))
    base_date = str(target.get('scheduled_date', '') or '').strip()
    if base_date:
        scheduled_dates.add(base_date)
    for scheduled_date in sorted(scheduled_dates):
        _delete_activity_occurrence_data(activity_id, scheduled_date)

    return True


def delete_activities_for_month(month: str) -> int:
    """Delete all activities whose `month` field matches YYYY-MM. Returns count deleted."""
    activities = get_activities()
    to_delete = [a for a in activities if a.get('month', '')[:7] == month[:7]]
    remaining = [a for a in activities if a.get('month', '')[:7] != month[:7]]
    _save(ACTIVITIES_FILE, remaining)
    for activity in to_delete:
        activity_id = activity.get('id', '')
        records = get_records()
        scheduled_dates = {
            str(r.get('date', '')).strip()
            for r in records
            if r.get('activity_id') == activity_id and str(r.get('date', '')).strip()
        }
        scheduled_dates.update(_excluded_dates(activity))
        base_date = str(activity.get('scheduled_date', '') or '').strip()
        if base_date:
            scheduled_dates.add(base_date)
        for scheduled_date in sorted(scheduled_dates):
            _delete_activity_occurrence_data(activity_id, scheduled_date)
    return len(to_delete)


def get_records() -> list:
    return _load(RECORDS_FILE)


def get_record(record_id: str) -> dict | None:
    return next((record for record in get_records() if record.get('id') == record_id), None)


def get_record_for_activity_date(activity_id: str, date: str) -> dict | None:
    matches = [
        record for record in get_records()
        if record.get('activity_id') == activity_id and record.get('date') == date
    ]
    if not matches:
        return None
    return sorted(matches, key=_record_sort_key, reverse=True)[0]


def start_record(activity_id: str, date: str, username: str, user_name: str) -> dict:
    existing = get_record_for_activity_date(activity_id, date)
    if existing:
        return existing

    activity = get_activity(activity_id)
    records = get_records()
    record = {
        'id': str(uuid.uuid4()),
        'activity_id': activity_id,
        'activity_name': activity.get('name', '') if activity else '',
        'date': date,
        'started_at': datetime.now().isoformat(),
        'started_by': username,
        'started_name': user_name,
        'excel_values': {},
        'before_photos': [],
        'after_photos': [],
        'completed': False,
        'completed_at': None,
    }
    records.append(record)
    _save(RECORDS_FILE, records)
    return record


def update_record(record_id: str, data: dict) -> dict | None:
    records = get_records()
    for record in records:
        if record.get('id') != record_id:
            continue
        record.update({key: value for key, value in data.items() if key != 'id'})
        _save(RECORDS_FILE, records)
        return record
    return None


def save_photo(record_id: str, phase: str, uploaded_file) -> str:
    dest_dir = PHOTOS_DIR / record_id / phase
    dest_dir.mkdir(parents=True, exist_ok=True)

    ext = Path(uploaded_file.name).suffix.lower() or '.jpg'
    filename = f'{uuid.uuid4().hex}{ext}'
    dest = dest_dir / filename
    with dest.open('wb') as handle:
        for chunk in uploaded_file.chunks():
            handle.write(chunk)

    rel = f'cmms/photos/{record_id}/{phase}/{filename}'
    record = get_record(record_id)
    if record:
        key = 'before_photos' if phase == 'before' else 'after_photos'
        photos = list(record.get(key, []))
        photos.append(rel)
        update_record(record_id, {key: photos})
    return rel


def delete_photo(record_id: str, phase: str, rel_path: str) -> bool:
    record = get_record(record_id)
    if not record:
        return False
    key = 'before_photos' if phase == 'before' else 'after_photos'
    photos = list(record.get(key, []))
    if rel_path not in photos:
        return False
    photos.remove(rel_path)
    update_record(record_id, {key: photos})
    full = MEDIA_ROOT / rel_path
    if full.exists():
        full.unlink()
    return True


CHECKLIST_DATA_FILE = ROOT_CHECKLISTS_DIR / 'checklist_data.xlsx'
CHECKLIST_DATA_SOURCE_URL = str(getattr(settings, 'CMMS_CHECKLIST_DATA_SOURCE_URL', '') or '').strip()
CHECKLIST_HTTP_TIMEOUT = int(getattr(settings, 'CMMS_CHECKLIST_HTTP_TIMEOUT', 10) or 10)


def _normalize_activity_name(value: str) -> str:
    return ' '.join(str(value or '').strip().lower().split())


def _clean_http_url(value: str) -> str:
    link = str(value or '').strip()
    if link.startswith('http://') or link.startswith('https://'):
        return link
    return ''


def _rows_to_checklist_activities(rows) -> list:
    activities_by_name: dict[str, dict] = {}
    for row_idx, row in enumerate(rows):
        if row_idx == 0:
            continue
        if not row or not row[0]:
            continue
        cells = [str(cell).strip() if cell is not None else '' for cell in row]
        name = cells[0]
        checklist_cell = cells[1] if len(cells) > 1 else ''
        permit_name = cells[2] if len(cells) > 2 else ''
        permit_link = _clean_http_url(cells[3] if len(cells) > 3 else '')
        report_name = cells[4] if len(cells) > 4 else ''
        report_link = _clean_http_url(cells[5] if len(cells) > 5 else '')
        checklist_link = _clean_http_url(checklist_cell)
        checklist_name = '' if checklist_link else checklist_cell

        if not report_link and _clean_http_url(report_name):
            report_link = _clean_http_url(report_name)
            report_name = ''
        if report_link and not report_name:
            report_name = 'CM Report'

        normalized = _normalize_activity_name(name)
        if not normalized:
            continue
        activity = activities_by_name.get(normalized)
        if not activity:
            activity = {
                'name': name,
                'link': checklist_link,
                'checklist_name': checklist_name,
                'checklist_link': checklist_link,
                'report_name': report_name,
                'report_link': report_link,
                'permits': [],
            }
            activities_by_name[normalized] = activity
        else:
            if not activity.get('link') and checklist_link:
                activity['link'] = checklist_link
            if not activity.get('checklist_link') and checklist_link:
                activity['checklist_link'] = checklist_link
            if not activity.get('checklist_name') and checklist_name:
                activity['checklist_name'] = checklist_name
            if not activity.get('report_name') and report_name:
                activity['report_name'] = report_name
            if not activity.get('report_link') and report_link:
                activity['report_link'] = report_link

        if permit_name or permit_link:
            normalized_permit = _normalize_activity_name(permit_name or permit_link)
            if normalized_permit:
                exists = any(
                    _normalize_activity_name(item.get('name', '')) == normalized_permit
                    and str(item.get('link', '') or '').strip() == permit_link
                    for item in activity['permits']
                )
                if not exists:
                    activity['permits'].append({
                        'name': permit_name or 'Permit',
                        'link': permit_link,
                    })
    return list(activities_by_name.values())


def _local_checklist_activities() -> list:
    if not CHECKLIST_DATA_FILE.exists():
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(CHECKLIST_DATA_FILE), data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        return _rows_to_checklist_activities(rows)
    except Exception:
        return []


def _build_google_sheet_csv_urls(sheet_url: str) -> list[str]:
    if not sheet_url:
        return []
    try:
        parsed = urlparse(sheet_url)
    except Exception:
        return []

    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9\-_]+)', parsed.path or '')
    if not match:
        return []

    sheet_id = match.group(1)
    query = parse_qs(parsed.query or '')
    gid = str(query.get('gid', [''])[0]).strip()

    urls = []
    export_params = {'format': 'csv'}
    if gid:
        export_params['gid'] = gid
    urls.append(f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?{urlencode(export_params)}')

    gviz_params = {'tqx': 'out:csv'}
    if gid:
        gviz_params['gid'] = gid
    urls.append(f'https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?{urlencode(gviz_params)}')
    return urls


def _remote_checklist_activities() -> list | None:
    source_reached = False
    for csv_url in _build_google_sheet_csv_urls(CHECKLIST_DATA_SOURCE_URL):
        try:
            request = Request(csv_url, headers={'User-Agent': 'HDEC-CMMS/1.0'})
            with urlopen(request, timeout=CHECKLIST_HTTP_TIMEOUT) as response:
                payload = response.read()
            text = payload.decode('utf-8-sig', errors='replace')
            if '<html' in text.lower():
                continue
            source_reached = True
            rows = list(csv.reader(io.StringIO(text)))
            activities = _rows_to_checklist_activities(rows)
            if activities:
                return activities
            if rows:
                return []
        except Exception:
            continue
    return None if not source_reached else []


def get_all_checklist_activities() -> list:
    """
    Return activity mappings from the live Google Sheet first, with
    Checklists/checklist_data.xlsx as a fallback when the remote source
    is unavailable.
    """
    now = datetime.now()
    cached_expires_at = _CHECKLIST_ACTIVITY_CACHE.get('expires_at')
    cached_activities = _CHECKLIST_ACTIVITY_CACHE.get('activities')
    if (
        isinstance(cached_expires_at, datetime)
        and cached_expires_at > now
        and isinstance(cached_activities, list)
    ):
        return cached_activities

    remote = _remote_checklist_activities()
    activities = remote if remote is not None else _local_checklist_activities()
    _CHECKLIST_ACTIVITY_CACHE['activities'] = activities
    _CHECKLIST_ACTIVITY_CACHE['expires_at'] = now + timedelta(seconds=CHECKLIST_ACTIVITY_CACHE_SECONDS)
    return activities


def _dedupe_permit_options(permits: list[dict]) -> list[dict]:
    result = []
    seen: set[tuple[str, str]] = set()
    for permit in permits or []:
        permit_name = str(permit.get('name', '') or '').strip()
        permit_link = str(permit.get('link', '') or '').strip()
        if not permit_name:
            continue
        clean_link = permit_link if permit_link.startswith('http') else ''
        key = (_normalize_activity_name(permit_name), clean_link)
        if key in seen:
            continue
        seen.add(key)
        result.append({
            'name': permit_name,
            'link': clean_link,
        })
    return result


def _match_activity_mapping(activity_name: str) -> dict | None:
    target = _normalize_activity_name(activity_name)
    if not target:
        return None

    exact = None
    best_partial = None
    best_score = 0
    for activity in get_all_checklist_activities():
        name = _normalize_activity_name(activity.get('name', ''))
        if not name:
            continue
        if name == target:
            exact = activity
            break
        if name in target or target in name:
            score = len(set(name.split()) & set(target.split()))
            if score > best_score:
                best_score = score
                best_partial = activity
    return exact or best_partial


def get_activity_permit_options(activity_name: str) -> list[dict]:
    """Return permit options configured for an activity from the live mapping sheet."""
    all_activities = get_all_checklist_activities()
    all_permits = _dedupe_permit_options([
        permit
        for activity in all_activities
        for permit in (activity.get('permits', []) or [])
    ])
    matched = _match_activity_mapping(activity_name)
    if not matched:
        return all_permits
    matched_permits = _dedupe_permit_options(matched.get('permits', []) or [])
    return matched_permits or all_permits


def get_checklist_link(activity_name: str) -> str | None:
    """
    Return the matching checklist link for the given activity name.
    The live Google Sheet is checked fresh each call, with the local
    checklist_data.xlsx file as fallback.
    """
    activity = _match_activity_mapping(activity_name) or {}
    link = _clean_http_url(activity.get('checklist_link') or activity.get('link'))
    return link or None


def get_checklist_name(activity_name: str) -> str:
    activity = _match_activity_mapping(activity_name) or {}
    label = str(activity.get('checklist_name', '') or '').strip()
    return label or 'Checklist'


def get_report_link(activity_name: str) -> str | None:
    activity = _match_activity_mapping(activity_name) or {}
    link = _clean_http_url(activity.get('report_link'))
    return link or None


def get_report_name(activity_name: str) -> str:
    activity = _match_activity_mapping(activity_name) or {}
    label = str(activity.get('report_name', '') or '').strip()
    return label or 'CM Report'


def get_activity_sheet_link(activity: dict | None) -> str:
    activity = activity or {}
    if _activity_type(activity.get('type', 'PM')) == 'CM':
        report_link = get_report_link(activity.get('name', ''))
        if report_link:
            return report_link
        if CMMS_CM_REPORT_TEMPLATE_URL.startswith('http'):
            return CMMS_CM_REPORT_TEMPLATE_URL
        return ''
    checklist_link = get_checklist_link(activity.get('name', ''))
    return checklist_link or ''


def get_activity_sheet_label(activity: dict | None) -> str:
    activity = activity or {}
    if _activity_type(activity.get('type', 'PM')) == 'CM':
        return get_report_name(activity.get('name', ''))
    return get_checklist_name(activity.get('name', ''))


def download_google_sheet_export(sheet_url: str, export_format: str = 'xlsx') -> bytes | None:
    link = str(sheet_url or '').strip()
    if not link:
        return None
    try:
        parsed = urlparse(link)
    except Exception:
        return None

    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9\-_]+)', parsed.path or '')
    if not match:
        return None

    query = parse_qs(parsed.query or '')
    fragment = parse_qs(str(parsed.fragment or '').lstrip('#'))
    gid = str(query.get('gid', [''])[0] or fragment.get('gid', [''])[0]).strip()
    params = {'format': export_format}
    if gid:
        params['gid'] = gid
    export_url = f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?{urlencode(params)}"

    try:
        request = Request(export_url, headers={'User-Agent': 'HDEC-CMMS/1.0'})
        with urlopen(request, timeout=CHECKLIST_HTTP_TIMEOUT) as response:
            payload = response.read()
        if not payload:
            return None
        if payload[:256].lower().startswith(b'<!doctype html') or payload[:256].lower().startswith(b'<html'):
            return None
        return payload
    except Exception:
        return None


def parse_excel_checklist(file_path):
    """
    Parse an Excel checklist file into structured JSON for web rendering.
    Uses pixel-perfect generic mode so the webpage shows the same layout as Excel.
    Skips helper sheets such as cover pages and workbook settings.
    Returns list of sheet dicts: {name, kind, column_widths, row_heights, cells}
    """
    return _parse_excel_checklist_generic(file_path)


def _excel_value_to_text(value) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        if value.time() == time(0, 0):
            return value.strftime('%Y-%m-%d')
        return value.strftime('%Y-%m-%d %H:%M')
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime('%H:%M')
    if isinstance(value, bool):
        return 'TRUE' if value else 'FALSE'
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _excel_color_to_css(color, *, transparent_is_none=True) -> str | None:
    if color is None:
        return None

    try:
        color_type = color.type
    except Exception:
        return None

    raw = None
    if color_type == 'rgb':
        try:
            raw = color.rgb
        except Exception:
            raw = None
    elif color_type == 'indexed':
        try:
            from openpyxl.styles.colors import COLOR_INDEX

            index = color.indexed
            if isinstance(index, int) and 0 <= index < len(COLOR_INDEX):
                raw = COLOR_INDEX[index]
        except Exception:
            raw = None

    if not isinstance(raw, str):
        return None

    raw = raw.strip().upper()
    if len(raw) == 8:
        alpha = raw[:2]
        rgb = raw[2:]
    elif len(raw) == 6:
        alpha = 'FF'
        rgb = raw
    else:
        return None

    if transparent_is_none and alpha == '00':
        return None
    if len(rgb) != 6:
        return None
    return f'#{rgb}'


def _excel_border_to_css(side) -> str | None:
    style = getattr(side, 'style', None)
    if not style:
        return None

    width_map = {
        'hair': '1px',
        'thin': '1px',
        'dotted': '1px',
        'dashed': '1px',
        'dashDot': '1px',
        'dashDotDot': '1px',
        'slantDashDot': '1px',
        'medium': '2px',
        'mediumDashed': '2px',
        'mediumDashDot': '2px',
        'mediumDashDotDot': '2px',
        'double': '3px',
        'thick': '3px',
    }
    style_map = {
        'double': 'double',
        'dashed': 'dashed',
        'mediumDashed': 'dashed',
        'dotted': 'dotted',
    }
    color = _excel_color_to_css(getattr(side, 'color', None), transparent_is_none=True) or '#111827'
    width = width_map.get(style, '1px')
    line_style = style_map.get(style, 'solid')
    return f'{width} {line_style} {color}'


def _excel_column_width_to_px(width) -> int:
    width = float(width or 8.43)
    if width <= 0:
        return 0
    return max(24, int(round(width * 7 + 5)))


def _excel_row_height_to_px(height) -> int:
    height = float(height or 15)
    if height <= 0:
        return 0
    return max(20, int(round(height * 4 / 3)))


def _excel_cell_style_css(cell) -> str:
    font = cell.font
    alignment = cell.alignment
    fill = cell.fill
    border = cell.border

    horizontal = (alignment.horizontal or '').lower()
    vertical = (alignment.vertical or '').lower()
    justify_map = {
        'center': 'center',
        'centercontinuous': 'center',
        'right': 'flex-end',
        'distributed': 'stretch',
        'justify': 'stretch',
    }
    align_map = {
        'center': 'center',
        'distributed': 'center',
        'justify': 'stretch',
        'bottom': 'flex-end',
    }
    style_bits = [
        'display:flex',
        f'justify-content:{justify_map.get(horizontal, "flex-start")}',
        f'align-items:{align_map.get(vertical, "flex-start")}',
        'padding:4px 6px',
        'overflow:hidden',
        'box-sizing:border-box',
    ]

    background = _excel_color_to_css(getattr(fill, 'fgColor', None), transparent_is_none=True)
    if background:
        style_bits.append(f'background:{background}')

    color = _excel_color_to_css(getattr(font, 'color', None), transparent_is_none=True)
    if color:
        style_bits.append(f'color:{color}')

    if font.name:
        family = str(font.name).replace('\\', '\\\\').replace("'", "\\'")
        style_bits.append(f"font-family:'{family}'")
    if font.sz:
        style_bits.append(f'font-size:{max(10, int(round(float(font.sz) * 4 / 3)))}px')
    if font.bold:
        style_bits.append('font-weight:700')
    if font.italic:
        style_bits.append('font-style:italic')
    if font.underline:
        style_bits.append('text-decoration:underline')

    text_align = {
        'center': 'center',
        'centercontinuous': 'center',
        'right': 'right',
        'fill': 'left',
        'justify': 'justify',
        'distributed': 'justify',
    }.get(horizontal, 'left')
    style_bits.append(f'text-align:{text_align}')

    white_space = 'pre-wrap' if alignment.wrap_text else 'nowrap'
    style_bits.append(f'white-space:{white_space}')

    border_top = _excel_border_to_css(border.top)
    border_right = _excel_border_to_css(border.right)
    border_bottom = _excel_border_to_css(border.bottom)
    border_left = _excel_border_to_css(border.left)
    if border_top:
        style_bits.append(f'border-top:{border_top}')
    if border_right:
        style_bits.append(f'border-right:{border_right}')
    if border_bottom:
        style_bits.append(f'border-bottom:{border_bottom}')
    if border_left:
        style_bits.append(f'border-left:{border_left}')

    return ';'.join(style_bits)


def _should_skip_checklist_sheet(worksheet) -> bool:
    title = re.sub(r'\s+', ' ', str(getattr(worksheet, 'title', '') or '')).strip().lower()
    helper_sheet_names = {
        'cover page',
        'setting',
        'settings',
        'google image',
        'google images',
    }
    return getattr(worksheet, 'sheet_state', 'visible') != 'visible' or title in helper_sheet_names


def _parse_excel_checklist_generic(file_path):
    import openpyxl
    from openpyxl.utils import get_column_letter, range_boundaries

    workbook = openpyxl.load_workbook(str(file_path), data_only=False)
    display_workbook = openpyxl.load_workbook(str(file_path), data_only=True)
    sheets = []

    for worksheet_index, worksheet in enumerate(workbook.worksheets):
        if _should_skip_checklist_sheet(worksheet):
            continue
        display_sheet = display_workbook.worksheets[worksheet_index]
        max_row = worksheet.max_row or 0
        max_col = worksheet.max_column or 0
        if max_row == 0 or max_col == 0:
            continue

        default_col_width = worksheet.sheet_format.defaultColWidth or 8.43
        default_row_height = worksheet.sheet_format.defaultRowHeight or 15
        column_widths = []
        row_heights = []
        for col_no in range(1, max_col + 1):
            column_letter = get_column_letter(col_no)
            dimension = worksheet.column_dimensions.get(column_letter)
            if dimension and dimension.hidden:
                column_widths.append(0)
            else:
                width = dimension.width if dimension and dimension.width is not None else default_col_width
                column_widths.append(_excel_column_width_to_px(width))

        for row_no in range(1, max_row + 1):
            dimension = worksheet.row_dimensions.get(row_no)
            if dimension and dimension.hidden:
                row_heights.append(0)
            else:
                height = dimension.height if dimension and dimension.height is not None else default_row_height
                row_heights.append(_excel_row_height_to_px(height))

        merge_map: dict[tuple[int, int], tuple[int, int]] = {}
        merged_children: set[tuple[int, int]] = set()
        for merged_range in worksheet.merged_cells.ranges:
            min_col, min_row, max_col_range, max_row_range = range_boundaries(str(merged_range))
            merge_map[(min_row, min_col)] = (
                max_row_range - min_row + 1,
                max_col_range - min_col + 1,
            )
            for row_no in range(min_row, max_row_range + 1):
                for col_no in range(min_col, max_col_range + 1):
                    if row_no == min_row and col_no == min_col:
                        continue
                    merged_children.add((row_no, col_no))

        cells = []
        for row_no in range(1, max_row + 1):
            if row_heights[row_no - 1] == 0:
                continue
            for col_no in range(1, max_col + 1):
                if column_widths[col_no - 1] == 0 or (row_no, col_no) in merged_children:
                    continue

                cell = worksheet.cell(row=row_no, column=col_no)
                display_cell = display_sheet[cell.coordinate]
                display_value = display_cell.value if cell.data_type == 'f' else cell.value
                value = _excel_value_to_text(display_value)
                rowspan, colspan = merge_map.get((row_no, col_no), (1, 1))
                effective_rowspan = len(
                    [height for height in row_heights[row_no - 1:row_no - 1 + rowspan] if height > 0]
                ) or 1
                effective_colspan = len(
                    [width for width in column_widths[col_no - 1:col_no - 1 + colspan] if width > 0]
                ) or 1
                style_css = _excel_cell_style_css(cell)
                if value or effective_rowspan > 1 or effective_colspan > 1 or style_css:
                    cells.append({
                        'row': row_no,
                        'col': col_no,
                        'ref': cell.coordinate,
                        'value': value,
                        'editable': cell.data_type != 'f',
                        'rowspan': effective_rowspan,
                        'colspan': effective_colspan,
                        'style': style_css,
                    })

        if cells:
            sheets.append({
                'name': worksheet.title,
                'kind': 'generic',
                'column_widths': column_widths,
                'row_heights': row_heights,
                'cells': cells,
            })

    return sheets


def fill_excel_checklist(original_path, sheet_values) -> io.BytesIO:
    """
    Write user-provided values back into a copy of the checklist.
    sheet_values: {sheet_name: {metadata:{key:val}, items:{str(sr_no):obs}, obs_rows:{str(num):text}}}
    Returns BytesIO of the filled workbook.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(str(original_path))

    for sheet_name, values in sheet_values.items():
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        if values.get('mode') == 'generic':
            for ref, value in values.get('cells', {}).items():
                worksheet[ref] = value
            continue
        meta_vals = values.get('metadata', {})
        item_vals = values.get('items', {})
        obs_vals = values.get('obs_rows', {})

        header_row = None
        obs_write_col = 16
        for row in worksheet.iter_rows():
            first = row[0].value
            if first and str(first).strip().lower() in ('sr. no.', 'sr.no.', 'sr no', 'block'):
                header_row = row[0].row
                for cell in row:
                    if cell.value and 'observation' in str(cell.value).lower():
                        obs_write_col = cell.column
                        break
                break

        for row in worksheet.iter_rows():
            first = row[0].value
            if not first:
                continue
            lower = str(first).lower()
            if 'due date' in lower:
                if 'due_date' in meta_vals:
                    worksheet.cell(row=row[0].row, column=3).value = meta_vals['due_date']
                if 'done_date' in meta_vals:
                    worksheet.cell(row=row[0].row, column=8).value = meta_vals['done_date']
            elif 'start time' in lower:
                if 'start_time' in meta_vals:
                    worksheet.cell(row=row[0].row, column=3).value = meta_vals['start_time']
                if 'completion_time' in meta_vals:
                    worksheet.cell(row=row[0].row, column=8).value = meta_vals['completion_time']
            elif 'technician' in lower:
                if 'technician' in meta_vals:
                    worksheet.cell(row=row[0].row, column=4).value = meta_vals['technician']
            elif 'supervisor' in lower:
                if 'supervisor' in meta_vals:
                    worksheet.cell(row=row[0].row, column=4).value = meta_vals['supervisor']
            elif 'maintenance engineer' in lower:
                if 'engineer' in meta_vals:
                    worksheet.cell(row=row[0].row, column=4).value = meta_vals['engineer']
            elif 'customer rep' in lower:
                if 'customer_rep' in meta_vals:
                    worksheet.cell(row=row[0].row, column=4).value = meta_vals['customer_rep']

        if header_row and item_vals:
            in_obs = False
            for row in worksheet.iter_rows(min_row=header_row + 1):
                first_cell = row[0]
                first_val = first_cell.value
                if first_val and str(first_val).lower().startswith('observation'):
                    in_obs = True
                    continue
                if first_val and str(first_val).lower().startswith('abbreviation'):
                    break
                if in_obs:
                    try:
                        num = int(first_val)
                        if str(num) in obs_vals:
                            worksheet.cell(row=first_cell.row, column=2).value = obs_vals[str(num)]
                    except (ValueError, TypeError):
                        pass
                    continue
                try:
                    sr_no = int(first_val) if first_val else None
                except (ValueError, TypeError):
                    sr_no = None
                if sr_no is not None and str(sr_no) in item_vals:
                    worksheet.cell(row=first_cell.row, column=obs_write_col).value = item_vals[str(sr_no)]

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def generate_zip(record_id: str) -> io.BytesIO | None:
    """
    Build ZIP:
      <activity>_<date>/
        <checklist>_completed.xlsx
        Before Pictures/
        After Pictures/
    """
    record = get_record(record_id)
    if not record:
        return None
    activity = get_activity(record.get('activity_id', '')) or {}

    folder = f"{record.get('activity_name', 'checklist')}_{record.get('date', '')}".replace(' ', '_')
    buffer = io.BytesIO()

    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as archive:
        checklist_path = resolve_checklist_path(activity.get('checklist_file', ''))
        linked_sheet_url = get_activity_sheet_link(activity)
        sheet_label = get_activity_sheet_label(activity)
        if checklist_path and checklist_path.exists():
            saved_values = record.get('excel_values', {})
            if saved_values:
                try:
                    workbook_buffer = fill_excel_checklist(checklist_path, saved_values)
                    archive.writestr(
                        f"{folder}/{checklist_path.stem}_completed{checklist_path.suffix}",
                        workbook_buffer.read(),
                    )
                except Exception:
                    archive.write(checklist_path, f"{folder}/{checklist_path.name}")
            else:
                archive.write(checklist_path, f"{folder}/{checklist_path.name}")
        elif linked_sheet_url:
            workbook_bytes = download_google_sheet_export(linked_sheet_url, 'xlsx')
            if workbook_bytes:
                safe_name = _slug(activity.get('name') or sheet_label or 'sheet')
                archive.writestr(
                    f"{folder}/{safe_name}_{sheet_label.lower().replace(' ', '_')}.xlsx",
                    workbook_bytes,
                )

        try:
            from .cmms_ptw_utils import annotate_permit, ensure_final_permit_pdf, get_permit_for_record

            permit = annotate_permit(get_permit_for_record(record_id))
        except Exception:
            permit = None
        if permit:
            permit_lines = [
                f"PTW Status: {permit.get('status_label', permit.get('status', ''))}",
                f"Permit Number: {permit.get('permit_number', '') or 'N/A'}",
                f"Isolation Certificate Number: {permit.get('isolation_cert_number', '') or 'N/A'}",
                f"Google Docs Permit Link: {permit.get('document_link', '') or 'N/A'}",
                f"Receiver: {permit.get('receiver_name', '') or 'N/A'}",
                f"Issuer: {permit.get('issuer_name', '') or 'N/A'}",
                f"HSE: {permit.get('hse_name', '') or 'N/A'}",
            ]
            archive.writestr(f"{folder}/PTW Details.txt", "\n".join(permit_lines))
            final_pdf = str(permit.get('final_pdf', '') or '').strip()
            if permit.get('status') == 'closed' and not final_pdf:
                try:
                    refreshed = annotate_permit(ensure_final_permit_pdf(permit))
                    final_pdf = str((refreshed or {}).get('final_pdf', '') or '').strip()
                except Exception:
                    final_pdf = ''
            if final_pdf:
                final_pdf_path = MEDIA_ROOT / final_pdf
                if final_pdf_path.exists():
                    archive.write(final_pdf_path, f"{folder}/{final_pdf_path.name}")

        for rel in record.get('before_photos', []):
            full = MEDIA_ROOT / rel
            if full.exists():
                archive.write(full, f"{folder}/Before Pictures/{full.name}")

        for rel in record.get('after_photos', []):
            full = MEDIA_ROOT / rel
            if full.exists():
                archive.write(full, f"{folder}/After Pictures/{full.name}")

    buffer.seek(0)
    return buffer


def get_checklist_files() -> list:
    """
    Return list of checklist workbooks available to the CMMS scheduler.
    The stored ``rel_path`` is a stable reference understood by
    ``resolve_checklist_path``.
    """
    files = []
    for path in _iter_checklist_paths():
        files.append({
            'name': path.name,
            'rel_path': _serialize_checklist_ref(path),
        })
    return files
