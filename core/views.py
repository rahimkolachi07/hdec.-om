import json, csv, io, urllib.request, urllib.parse
from collections import Counter
from datetime import datetime
from functools import wraps
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .auth_utils import (
    authenticate, get_all_users, create_user, delete_user, change_password,
    update_user_permissions, MODULES, DEFAULT_PERMISSIONS, has_permission,
    get_user_detail, normalize_user_state,
    filter_countries_for_user, filter_projects_for_user,
    can_access_country, can_access_project,
)
from .project_utils import (
    get_countries, get_country, get_project, get_project_module_cards,
    create_country, update_country, delete_country,
    create_project, update_project, delete_project,
    MODULE_META, ALL_MODULE_IDS,
    CATEGORY_META, ALL_CATEGORY_IDS,
    get_project_categories, get_all_modules_flat,
)

# ── HELPERS ────────────────────────────────────────────────────────────────

def get_user(request):
    user = normalize_user_state(request.session.get('hdec_user'))
    if user != request.session.get('hdec_user'):
        request.session['hdec_user'] = user
        request.session.modified = True
    return user

def is_admin(request):
    u = get_user(request)
    return bool(u and u.get('role') == 'admin')

def _ctx(request, extra=None):
    ctx = {
        'current_user': get_user(request) or {},
        'is_admin': is_admin(request),
    }
    if extra:
        ctx.update(extra)
    return ctx

def login_required(fn):
    @wraps(fn)
    def wrapper(request, *args, **kwargs):
        if not get_user(request):
            return redirect('/login/')
        return fn(request, *args, **kwargs)
    return wrapper


def _has_any_project_access(user):
    if not user:
        return False
    if user.get('role') == 'admin':
        return True
    for country in get_countries():
        country_id = country.get('id', '')
        if not can_access_country(user, country_id):
            continue
        for project in country.get('projects', []):
            if can_access_project(user, country_id, project.get('id', '')):
                return True
    return False


def _require_module_page(request, module_id: str, level: str = 'view'):
    user = get_user(request)
    if not user:
        return redirect('/login/')
    if not _has_any_project_access(user):
        return redirect('/')
    if not has_permission(user, module_id, level):
        return redirect('/')
    return None


def _require_module_api(request, module_id: str, level: str = 'view'):
    user = get_user(request)
    if not user:
        return None, JsonResponse({'error': 'Login required'}, status=401)
    if not _has_any_project_access(user) or not has_permission(user, module_id, level):
        return None, JsonResponse({'error': 'Forbidden'}, status=403)
    return user, None

# ── AUTH VIEWS ─────────────────────────────────────────────────────────────

def login_view(request):
    # Already logged in → go to dashboard
    if get_user(request):
        return redirect('/')

    error = None
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(username, password)
        if user:
            request.session['hdec_user'] = user
            request.session.modified = True
            return redirect('/')
        error = 'Invalid username or password. Please try again.'

    return render(request, 'core/login.html', {'error': error})


def logout_view(request):
    request.session.flush()
    return redirect('/login/')


@login_required
def admin_panel(request):
    if not is_admin(request):
        return redirect('/')
    users = get_all_users()
    countries = get_countries()
    countries_ctx = [{
        'id': c.get('id', ''),
        'name': c.get('name', ''),
        'flag': c.get('flag', ''),
        'projects': [
            {'id': p.get('id', ''), 'name': p.get('name', '')}
            for p in c.get('projects', [])
        ],
    } for c in countries]
    return render(request, 'core/admin.html', {
        **_ctx(request),
        'users_json': json.dumps(users),
        'modules_json': json.dumps(MODULES),
        'defaults_json': json.dumps(DEFAULT_PERMISSIONS),
        'countries_json': json.dumps(countries_ctx),
    })


@csrf_exempt
def admin_api(request):
    user = get_user(request)
    if not user or user.get('role') != 'admin':
        return JsonResponse({'error': 'Forbidden'}, status=403)
    if request.method == 'POST':
        data = json.loads(request.body)
        action = data.get('action')
        if action == 'create':
            ok, msg = create_user(
                data.get('username', ''), data.get('password', ''),
                data.get('name', ''), data.get('role', 'viewer'),
                data.get('email', ''), data.get('permissions'), data.get('access')
            )
            return JsonResponse({'ok': ok, 'msg': msg})
        elif action == 'delete':
            ok, msg = delete_user(data.get('username', ''))
            return JsonResponse({'ok': ok, 'msg': msg})
        elif action == 'change_password':
            ok, msg = change_password(data.get('username', ''), data.get('password', ''))
            return JsonResponse({'ok': ok, 'msg': msg})
        elif action == 'update_permissions':
            username = data.get('username', '')
            ok, msg = update_user_permissions(username, data.get('permissions', {}), data.get('access'))
            current = get_user(request)
            if ok and current and current.get('username') == username.strip().lower():
                request.session['hdec_user'] = get_user_detail(username)
                request.session.modified = True
            return JsonResponse({'ok': ok, 'msg': msg})
        elif action == 'get_defaults':
            role = data.get('role', 'viewer')
            return JsonResponse({'ok': True, 'permissions': DEFAULT_PERMISSIONS.get(role, {}), 'access': {'overall': 'all', 'countries': [], 'projects': []}})
    return JsonResponse({'error': 'Bad request'}, status=400)

# ── SHEET CONFIG ───────────────────────────────────────────────────────────

SHEET_ID = "1EKrRePyskWHJOPIljD7GCGAvyO4B9OMz80o8X4Cr1Ik"

TRACING_SHEETS = [
    {"slug": "statistics-pm",        "name": "Statistics PM",         "icon": "📊", "color": "#3b9eff",  "sheet": "Statistics PM"},
    {"slug": "statistics-cm",        "name": "Statistics CM",         "icon": "📈", "color": "#00e5c8",  "sheet": "Statistics CM"},
    {"slug": "pm-tracing",           "name": "PM Tracing",            "icon": "🔧", "color": "#a259ff",  "sheet": "PM Tracing"},
    {"slug": "scb",                  "name": "SCB",                   "icon": "⚡", "color": "#f0c040",  "sheet": "SCB"},
    {"slug": "trackers",             "name": "Trackers",              "icon": "📡", "color": "#22c55e",  "sheet": "Trackers"},
    {"slug": "inverter-mvps",        "name": "Inverter & MVPS",       "icon": "🔌", "color": "#ff6b35",  "sheet": "Inverter and MVPS"},
    {"slug": "strings-pv-modules",   "name": "Strings & PV Modules",  "icon": "☀️", "color": "#e879f9",  "sheet": "Strings & PV Modules"},
    {"slug": "pv-equipment-failure", "name": "PV Equipment Failure",  "icon": "⚠️", "color": "#ef4444",  "sheet": "PV Equipments Failure"},
    {"slug": "ss-equipment-failure", "name": "SS Equipment Failure",  "icon": "🔴", "color": "#fb923c",  "sheet": "SS Equipments failure"},
    {"slug": "ss-observations",      "name": "SS Observations",       "icon": "🔍", "color": "#38bdf8",  "sheet": "SS Observations"},
    {"slug": "pv-observations",      "name": "PV Observations",       "icon": "🌞", "color": "#4ade80",  "sheet": "PV Observations"},
]
SLUG_MAP = {s["slug"]: s for s in TRACING_SHEETS}


def fetch_sheet_csv(sheet_name):
    url = (f"https://docs.google.com/spreadsheets/d/{SHEET_ID}"
           f"/export?format=csv&sheet={urllib.parse.quote(sheet_name)}")
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as r:
            return r.read().decode("utf-8"), None
    except Exception as e:
        return None, str(e)


def parse_generic_sheet(raw_csv):
    reader = csv.reader(io.StringIO(raw_csv))
    rows = list(reader)
    if not rows:
        return {"headers": [], "records": [], "aggregates": {}, "summary_kv": {}}
    header_idx, summary_kv = 0, {}
    for i, row in enumerate(rows[:8]):
        non_empty = [c for c in row if c.strip()]
        if len(non_empty) >= 3:
            if any(c.replace(" ", "").isalpha() or len(c) > 3
                   for c in [c.strip() for c in row[:4] if c.strip()]):
                header_idx = i
                break
        elif non_empty:
            k = row[0].strip()
            v = row[2].strip() if len(row) > 2 else (row[1].strip() if len(row) > 1 else "")
            if k:
                summary_kv[k] = v
    headers = [h.strip() for h in rows[header_idx]]
    while headers and not headers[-1]:
        headers.pop()
    records = []
    for row in rows[header_idx + 1:]:
        if not any(c.strip() for c in row):
            continue
        records.append({headers[j]: (row[j].strip() if j < len(row) else "")
                        for j in range(len(headers))})
    aggregates = {}
    for h in headers:
        if not h:
            continue
        vals = [r[h] for r in records if r.get(h)]
        if 1 < len(set(vals)) <= 25:
            aggregates[h] = dict(Counter(vals).most_common(20))
    return {"headers": headers, "records": records, "aggregates": aggregates,
            "summary_kv": summary_kv, "total_records": len(records),
            "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

# ── PAGE VIEWS ─────────────────────────────────────────────────────────────

@login_required
def home(request):
    """Landing page — country selection."""
    user = get_user(request)
    countries = filter_countries_for_user(user, get_countries())
    countries_ctx = [
        {
            **c,
            'project_count': len(filter_projects_for_user(user, c.get('id', ''), c.get('projects', []))),
            'projects': [],
        }
        for c in countries
    ]
    return render(request, 'core/landing.html', _ctx(request, {
        'countries': countries_ctx,
    }))


@login_required
def country_view(request, country_id):
    """Projects list for a specific country."""
    user = get_user(request)
    if not can_access_country(user, country_id):
        return redirect('/')
    country = get_country(country_id)
    if not country:
        return redirect('/')

    module_icons = {mid: meta['icon'] for mid, meta in MODULE_META.items()}

    projects_ctx = []
    for p in filter_projects_for_user(user, country_id, country.get('projects', [])):
        all_mods = get_all_modules_flat(p)
        cats = get_project_categories(p)
        cats_json = {
            cid: {'modules': cats.get(cid, {}).get('modules', [])}
            for cid in ALL_CATEGORY_IDS
        }
        projects_ctx.append({
            **p,
            'module_count': len(all_mods),
            'module_icons': [module_icons.get(m, '📦') for m in all_mods[:6]],
            'categories_json': json.dumps(cats_json),
            'created_at': p.get('created_at', '')[:10],
        })

    all_modules_json = json.dumps([
        {'id': mid, 'label': meta['label'], 'icon': meta['icon']}
        for mid, meta in MODULE_META.items()
    ])
    all_categories_json = json.dumps([
        {'id': cid, 'label': meta['label'], 'icon': meta['icon'], 'color': meta['color']}
        for cid, meta in CATEGORY_META.items()
    ])

    return render(request, 'core/country_projects.html', _ctx(request, {
        'country': country,
        'projects': projects_ctx,
        'all_modules_json': all_modules_json,
        'all_categories_json': all_categories_json,
    }))


@login_required
def project_hub_view(request, country_id, project_id):
    """Category selector hub for a specific project."""
    user = get_user(request)
    if not can_access_project(user, country_id, project_id):
        return redirect(f'/c/{country_id}/' if can_access_country(user, country_id) else '/')
    country = get_country(country_id)
    project = get_project(country_id, project_id)
    if not country or not project:
        return redirect('/')

    cats = get_project_categories(project)
    category_cards = []
    for cid in ALL_CATEGORY_IDS:
        meta = CATEGORY_META[cid]
        mods = cats.get(cid, {}).get('modules', [])
        preview_icons = [MODULE_META[m]['icon'] for m in mods if m in MODULE_META][:5]
        category_cards.append({
            'id': cid,
            'label': meta['label'],
            'icon': meta['icon'],
            'color': meta['color'],
            'color_rgb': meta['color_rgb'],
            'desc': meta['desc'],
            'module_count': len(mods),
            'module_icons': preview_icons,
            'remaining_count': max(len(mods) - len(preview_icons), 0),
        })

    return render(request, 'core/project_hub.html', _ctx(request, {
        'country': country,
        'project': project,
        'category_cards': category_cards,
    }))


@login_required
def category_hub_view(request, country_id, project_id, category):
    """Module hub for a specific category within a project."""
    user = get_user(request)
    if not can_access_project(user, country_id, project_id):
        return redirect(f'/c/{country_id}/' if can_access_country(user, country_id) else '/')
    country = get_country(country_id)
    project = get_project(country_id, project_id)
    if not country or not project:
        return redirect('/')
    if category not in CATEGORY_META:
        return redirect(f'/p/{country_id}/{project_id}/')

    permissions = user.get('permissions', {}) if user else {}

    def hex_to_rgb(hex_color):
        h = hex_color.lstrip('#')
        try:
            r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
            return f'{r},{g},{b}'
        except Exception:
            return '59,158,255'

    all_cards = get_project_module_cards(project, permissions, country_id, project_id, category)
    for card in all_cards:
        card['color_rgb'] = hex_to_rgb(card['color'])

    cmms_ids = {'activities', 'permits', 'handover'}
    cmms_cards = [c for c in all_cards if c['id'] in cmms_ids]
    other_cards = [c for c in all_cards if c['id'] not in cmms_ids]

    cat_meta = CATEGORY_META[category]

    return render(request, 'core/category_hub.html', _ctx(request, {
        'country': country,
        'project': project,
        'category': category,
        'cat_meta': cat_meta,
        'module_cards': all_cards,
        'cmms_cards': cmms_cards,
        'other_cards': other_cards,
    }))


@login_required
def legacy_store_redirect(request):
    """Redirect old/global Store entry points to the first configured project store."""
    user = get_user(request)
    countries = get_countries()
    for country in countries:
        if not can_access_country(user, country.get('id', '')):
            continue
        for project in country.get('projects', []):
            if not can_access_project(user, country.get('id', ''), project.get('id', '')):
                continue
            categories = get_project_categories(project)
            for category_id in ALL_CATEGORY_IDS:
                modules = categories.get(category_id, {}).get('modules', [])
                if 'store' in modules:
                    return redirect(f"/p/{country['id']}/{project['id']}/{category_id}/store/")
    return redirect('/')


@csrf_exempt
def projects_api(request):
    """Admin API for managing countries and projects."""
    user = get_user(request)
    if not user or user.get('role') != 'admin':
        return JsonResponse({'error': 'Forbidden'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    action = data.get('action', '')

    if action == 'add_country':
        name = data.get('name', '').strip()
        if not name:
            return JsonResponse({'error': 'Name is required'})
        new_id = create_country(name, data.get('flag', '🏴'), data.get('color', '#3b9eff'))
        return JsonResponse({'ok': True, 'id': new_id})

    elif action == 'update_country':
        cid = data.get('id', '')
        if not cid:
            return JsonResponse({'error': 'id is required'})
        update_country(cid, {
            'name': data.get('name', '').strip(),
            'flag': data.get('flag', '🏴'),
            'color': data.get('color', '#3b9eff'),
        })
        return JsonResponse({'ok': True})

    elif action == 'delete_country':
        cid = data.get('id', '')
        if not cid:
            return JsonResponse({'error': 'id is required'})
        delete_country(cid)
        return JsonResponse({'ok': True})

    elif action == 'add_project':
        cid = data.get('country_id', '')
        name = data.get('name', '').strip()
        if not cid or not name:
            return JsonResponse({'error': 'country_id and name are required'})
        # Accept either new `categories` dict or legacy `modules` list
        if 'categories' in data:
            categories = data['categories']
        else:
            # Legacy: all selected modules go to maintenance
            modules = data.get('modules', ALL_MODULE_IDS[:])
            categories = {cid2: {'modules': []} for cid2 in ALL_CATEGORY_IDS}
            categories['maintenance']['modules'] = modules
        new_id = create_project(cid, name, data.get('description', ''), categories)
        if not new_id:
            return JsonResponse({'error': 'Country not found'})
        return JsonResponse({'ok': True, 'id': new_id})

    elif action == 'update_project':
        cid = data.get('country_id', '')
        pid = data.get('id', '')
        if not cid or not pid:
            return JsonResponse({'error': 'country_id and id are required'})
        fields = {}
        if 'name' in data:
            fields['name'] = data['name'].strip()
        if 'description' in data:
            fields['description'] = data['description'].strip()
        if 'categories' in data:
            cats = {}
            for cat_id in ALL_CATEGORY_IDS:
                cat_data = data['categories'].get(cat_id, {})
                cats[cat_id] = {
                    'modules': [m for m in cat_data.get('modules', []) if m in MODULE_META]
                }
            fields['categories'] = cats
        update_project(cid, pid, fields)
        return JsonResponse({'ok': True})

    elif action == 'delete_project':
        cid = data.get('country_id', '')
        pid = data.get('id', '')
        if not cid or not pid:
            return JsonResponse({'error': 'country_id and id are required'})
        delete_project(cid, pid)
        return JsonResponse({'ok': True})

    return JsonResponse({'error': f'Unknown action: {action}'}, status=400)

@login_required
def manpower(request):
    redir = _require_module_page(request, 'manpower', 'view')
    if redir:
        return redir
    import json as _j
    from pathlib import Path as _P
    sf = _P(__file__).resolve().parent.parent / 'schedule_store.json'
    stored = None
    if sf.exists():
        try:
            with open(sf) as f:
                stored = _j.load(f)
        except Exception:
            pass

    ctx = _ctx(request)
    if stored:
        ctx['engineers_json']   = _j.dumps(stored.get('engineers', []))
        ctx['technicians_json'] = _j.dumps(stored.get('technicians', []))
        # Derive sorted date lists from the stored data
        eng_dates = sorted(set(
            d for p in stored.get('engineers', [])
            for d in p.get('schedule', {})
        ))
        tech_dates = sorted(set(
            d for p in stored.get('technicians', [])
            for d in p.get('schedule', {})
        ))
        ctx['eng_dates_json']   = _j.dumps(eng_dates)
        ctx['tech_dates_json']  = _j.dumps(tech_dates)
        ctx['source_file']      = stored.get('source_file', '')
        ctx['imported_at']      = stored.get('imported_at', '')
        ctx['data_source']      = 'imported'
    else:
        # No import yet — use the original Excel data as default
        import json as _jd
        _default_eng  = [{"dept": "Maintenance Team", "role": "Maintenance team leader", "name": "Ahmed israr", "schedule": {"2026-03-19": "ON leave"}}, {"dept": "Maintenance Team", "role": "Maintenance team leader", "name": "Khalid ali", "schedule": {"2026-03-19": "ON leave"}}, {"dept": "Maintenance Team", "role": "Maintenance team engineer", "name": "Farooq ahmed", "schedule": {"2026-03-19": "ON leave", "2026-04-05": "Day", "2026-04-06": "Night", "2026-04-07": "Rest", "2026-04-08": "Rest", "2026-04-09": "Day", "2026-04-10": "Night", "2026-04-11": "Rest", "2026-04-12": "Rest", "2026-04-13": "Day", "2026-04-14": "Night", "2026-04-15": "Rest", "2026-04-16": "Rest", "2026-04-17": "Day", "2026-04-18": "Night", "2026-04-19": "Rest", "2026-04-20": "Rest", "2026-04-21": "Day", "2026-04-22": "Night", "2026-04-23": "Rest", "2026-04-24": "Rest", "2026-04-25": "Day", "2026-04-26": "Night", "2026-04-27": "Rest", "2026-04-28": "Rest", "2026-04-29": "Day", "2026-04-30": "Night"}}, {"dept": "Maintenance Team", "role": "Maintenance team engineer", "name": "Abdul Rahim", "schedule": {"2026-03-19": "EID leave", "2026-03-20": "OFF", "2026-03-21": "Day", "2026-03-22": "Day", "2026-03-23": "Day", "2026-03-24": "Day", "2026-03-25": "Day", "2026-03-26": "Day", "2026-03-27": "OFF", "2026-03-28": "EID leave", "2026-03-29": "General", "2026-03-30": "General", "2026-03-31": "General", "2026-04-01": "General", "2026-04-02": "General", "2026-04-03": "OFF", "2026-04-04": "General", "2026-04-05": "General", "2026-04-06": "General", "2026-04-07": "General", "2026-04-08": "General", "2026-04-09": "General", "2026-04-10": "OFF", "2026-04-11": "General", "2026-04-12": "General", "2026-04-13": "General", "2026-04-14": "General", "2026-04-15": "General", "2026-04-16": "General", "2026-04-17": "OFF", "2026-04-18": "General", "2026-04-19": "General", "2026-04-20": "General", "2026-04-21": "General", "2026-04-22": "General", "2026-04-23": "General", "2026-04-24": "OFF", "2026-04-25": "General", "2026-04-26": "General", "2026-04-27": "General", "2026-04-28": "General", "2026-04-29": "General", "2026-04-30": "General"}}, {"dept": "Maintenance Team", "role": "Maintenance team engineer", "name": "Sheeraz", "schedule": {"2026-03-19": "Day", "2026-03-20": "Rest", "2026-03-21": "EID leave", "2026-03-22": "Night", "2026-03-23": "Rest", "2026-03-24": "Rest", "2026-03-25": "EID leave", "2026-03-26": "Night", "2026-03-27": "Rest", "2026-03-28": "Rest", "2026-03-29": "Day", "2026-03-30": "Night", "2026-03-31": "Rest", "2026-04-01": "Rest", "2026-04-02": "Day", "2026-04-03": "Night", "2026-04-04": "Rest", "2026-04-05": "Rest", "2026-04-06": "Day", "2026-04-07": "Night", "2026-04-08": "Rest", "2026-04-09": "Rest", "2026-04-10": "Day", "2026-04-11": "Night", "2026-04-12": "Rest", "2026-04-13": "Rest", "2026-04-14": "Day", "2026-04-15": "Night", "2026-04-16": "Rest", "2026-04-17": "Rest", "2026-04-18": "Day", "2026-04-19": "Night", "2026-04-20": "Rest", "2026-04-21": "Rest", "2026-04-22": "Day", "2026-04-23": "Night", "2026-04-24": "Rest", "2026-04-25": "Rest", "2026-04-26": "Day", "2026-04-27": "Night", "2026-04-28": "Rest", "2026-04-29": "Rest", "2026-04-30": "Day"}}, {"dept": "Maintenance Team", "role": "Maintenance Engineer", "name": "Muhammad Naveed", "schedule": {"2026-03-19": "EID leave", "2026-03-20": "OFF", "2026-03-21": "EID leave", "2026-03-22": "EID leave", "2026-03-23": "General", "2026-03-24": "General", "2026-03-25": "General", "2026-03-26": "General", "2026-03-27": "OFF", "2026-03-28": "General", "2026-03-29": "General", "2026-03-30": "General", "2026-03-31": "General", "2026-04-01": "General", "2026-04-02": "General", "2026-04-03": "OFF", "2026-04-04": "General", "2026-04-05": "General", "2026-04-06": "General", "2026-04-07": "General", "2026-04-08": "General", "2026-04-09": "General", "2026-04-10": "OFF", "2026-04-11": "General", "2026-04-12": "General", "2026-04-13": "General", "2026-04-14": "General", "2026-04-15": "General", "2026-04-16": "General", "2026-04-17": "OFF", "2026-04-18": "General", "2026-04-19": "General", "2026-04-20": "General", "2026-04-21": "General", "2026-04-22": "General", "2026-04-23": "General", "2026-04-24": "OFF", "2026-04-25": "General", "2026-04-26": "General", "2026-04-27": "General", "2026-04-28": "General", "2026-04-29": "General", "2026-04-30": "General"}}, {"dept": "Maintenance Team", "role": "Maintenance Engineer", "name": "Ali Raza", "schedule": {"2026-03-19": "Night", "2026-03-20": "Rest", "2026-03-21": "Rest", "2026-03-22": "EID leave", "2026-03-23": "Night", "2026-03-24": "Rest", "2026-03-25": "Rest", "2026-03-26": "EID leave", "2026-03-27": "Night", "2026-03-28": "Rest", "2026-03-29": "Rest", "2026-03-30": "Day", "2026-03-31": "Night", "2026-04-01": "Rest", "2026-04-02": "Rest", "2026-04-03": "Day", "2026-04-04": "Night", "2026-04-05": "Rest", "2026-04-06": "Rest", "2026-04-07": "Day", "2026-04-08": "Night", "2026-04-09": "Rest", "2026-04-10": "Rest", "2026-04-11": "Day", "2026-04-12": "Night", "2026-04-13": "Rest", "2026-04-14": "Rest", "2026-04-15": "Day", "2026-04-16": "Night", "2026-04-17": "Rest", "2026-04-18": "Rest", "2026-04-19": "Day", "2026-04-20": "Night", "2026-04-21": "Rest", "2026-04-22": "Rest", "2026-04-23": "Day", "2026-04-24": "Night", "2026-04-25": "Rest", "2026-04-26": "Rest", "2026-04-27": "Day", "2026-04-28": "Night", "2026-04-29": "Rest", "2026-04-30": "Rest"}}, {"dept": "Maintenance Team", "role": "Maintenance Engineer", "name": "Abdullah", "schedule": {"2026-03-19": "EID leave", "2026-03-20": "OFF", "2026-03-21": "EID leave", "2026-03-22": "EID leave", "2026-03-23": "General", "2026-03-24": "General", "2026-03-25": "General", "2026-03-26": "General", "2026-03-27": "OFF", "2026-03-28": "General", "2026-03-29": "General", "2026-03-30": "General", "2026-03-31": "General", "2026-04-01": "General", "2026-04-02": "General", "2026-04-03": "OFF", "2026-04-04": "General", "2026-04-05": "General", "2026-04-06": "General", "2026-04-07": "General", "2026-04-08": "General", "2026-04-09": "General", "2026-04-10": "OFF", "2026-04-11": "General", "2026-04-12": "General", "2026-04-13": "General", "2026-04-14": "General", "2026-04-15": "General", "2026-04-16": "General", "2026-04-17": "OFF", "2026-04-18": "General", "2026-04-19": "General", "2026-04-20": "General", "2026-04-21": "General", "2026-04-22": "General", "2026-04-23": "General", "2026-04-24": "OFF", "2026-04-25": "General", "2026-04-26": "General", "2026-04-27": "General", "2026-04-28": "General", "2026-04-29": "General", "2026-04-30": "General"}}, {"dept": "Maintenance Team", "role": "Maintenance Engineer", "name": "Basit", "schedule": {"2026-03-19": "Rest", "2026-03-20": "Day", "2026-03-21": "Night", "2026-03-22": "Rest", "2026-03-23": "Rest", "2026-03-24": "EID leave", "2026-03-25": "Night", "2026-03-26": "Rest", "2026-03-27": "Rest", "2026-03-28": "Day", "2026-03-29": "Night", "2026-03-30": "Rest", "2026-03-31": "Rest", "2026-04-01": "Day", "2026-04-02": "Night", "2026-04-03": "Rest", "2026-04-04": "Rest", "2026-04-05": "Day", "2026-04-06": "Night", "2026-04-07": "Rest", "2026-04-08": "Day", "2026-04-09": "Night", "2026-04-10": "Rest", "2026-04-11": "Rest", "2026-04-12": "Day", "2026-04-13": "Night", "2026-04-14": "Rest", "2026-04-15": "Rest", "2026-04-16": "Day", "2026-04-17": "Night", "2026-04-18": "Rest", "2026-04-19": "Rest", "2026-04-20": "Day", "2026-04-21": "Night", "2026-04-22": "Rest", "2026-04-23": "Rest", "2026-04-24": "Day", "2026-04-25": "Night", "2026-04-26": "Rest", "2026-04-27": "Rest", "2026-04-28": "Day", "2026-04-29": "Night", "2026-04-30": "Rest"}}, {"dept": "Maintenance Team", "role": "Maintenance Engineer", "name": "Adnan", "schedule": {"2026-03-19": "Rest", "2026-03-20": "Night", "2026-03-21": "Rest", "2026-03-22": "Rest", "2026-03-23": "EID leave", "2026-03-24": "Night", "2026-03-25": "Rest", "2026-03-26": "Rest", "2026-03-27": "Day", "2026-03-28": "Night", "2026-03-29": "Rest", "2026-03-30": "Rest", "2026-03-31": "Day", "2026-04-01": "Night", "2026-04-02": "Rest", "2026-04-03": "Rest", "2026-04-04": "Day", "2026-04-05": "Night", "2026-04-06": "Rest", "2026-04-07": "Rest", "2026-04-08": "EID leave", "2026-04-09": "EID leave", "2026-04-10": "Rest", "2026-04-11": "Day", "2026-04-12": "Night", "2026-04-13": "Rest", "2026-04-14": "Rest", "2026-04-15": "Day", "2026-04-16": "Night", "2026-04-17": "Rest", "2026-04-18": "Rest", "2026-04-19": "Day", "2026-04-20": "Night", "2026-04-21": "Rest", "2026-04-22": "Rest", "2026-04-23": "Day", "2026-04-24": "Night", "2026-04-25": "Rest", "2026-04-26": "Rest", "2026-04-27": "Day", "2026-04-28": "Night", "2026-04-29": "Rest", "2026-04-30": "Rest"}}]
        _default_tech = [{"name": "Zubair", "role": "Technician", "schedule": {"2026-03-18": "Day", "2026-03-19": "Day", "2026-03-20": "Day", "2026-03-21": "Day", "2026-03-22": "Day", "2026-03-23": "Eid Leave", "2026-03-27": "Off", "2026-03-28": "Day", "2026-03-29": "Day", "2026-03-30": "Day", "2026-03-31": "Day", "2026-04-01": "Day", "2026-04-02": "Day", "2026-04-03": "Off", "2026-04-04": "Night", "2026-04-05": "Night", "2026-04-06": "Night", "2026-04-07": "Night", "2026-04-08": "Night", "2026-04-09": "Night", "2026-04-10": "Off", "2026-04-11": "Day", "2026-04-12": "Day", "2026-04-13": "Day", "2026-04-14": "Day", "2026-04-15": "Day", "2026-04-16": "Day", "2026-04-17": "Off", "2026-04-18": "Day", "2026-04-19": "Day", "2026-04-20": "Day"}}, {"name": "Nurul", "role": "Technician", "schedule": {"2026-03-18": "Night", "2026-03-19": "Eid Leave", "2026-03-23": "Day", "2026-03-24": "Day", "2026-03-25": "Day", "2026-03-26": "Day", "2026-03-27": "Off", "2026-03-28": "Day", "2026-03-29": "Day", "2026-03-30": "Day", "2026-03-31": "Day", "2026-04-01": "Day", "2026-04-02": "Day", "2026-04-03": "Off", "2026-04-04": "Day", "2026-04-05": "Day", "2026-04-06": "Day", "2026-04-07": "Day", "2026-04-08": "Day", "2026-04-09": "Day", "2026-04-10": "Off", "2026-04-11": "Night", "2026-04-12": "Night", "2026-04-13": "Night", "2026-04-14": "Night", "2026-04-15": "Night", "2026-04-16": "Night", "2026-04-17": "Off", "2026-04-18": "Day", "2026-04-19": "Day", "2026-04-20": "Day"}}, {"name": "Shabbir", "role": "Technician", "schedule": {"2026-03-18": "Night", "2026-03-19": "Eid Leave", "2026-03-23": "Day", "2026-03-24": "Day", "2026-03-25": "Day", "2026-03-26": "Day", "2026-03-27": "Off", "2026-03-28": "Night", "2026-03-29": "Night", "2026-03-30": "Night", "2026-03-31": "Night", "2026-04-01": "Night", "2026-04-02": "Night", "2026-04-03": "Off", "2026-04-04": "Day", "2026-04-05": "Day", "2026-04-06": "Day", "2026-04-07": "Day", "2026-04-08": "Day", "2026-04-09": "Day", "2026-04-10": "Off", "2026-04-11": "Day", "2026-04-12": "Day", "2026-04-13": "Day", "2026-04-14": "Day", "2026-04-15": "Day", "2026-04-16": "Day", "2026-04-17": "Off", "2026-04-18": "Night", "2026-04-19": "Night", "2026-04-20": "Night"}}, {"name": "Qadeer", "role": "Technician", "schedule": {"2026-03-18": "Day", "2026-03-19": "Eid Leave", "2026-03-23": "Day", "2026-03-24": "Day", "2026-03-25": "Day", "2026-03-26": "Day", "2026-03-27": "Day", "2026-03-28": "Off", "2026-03-29": "Day", "2026-03-30": "Day", "2026-03-31": "Day", "2026-04-01": "Day", "2026-04-02": "Day", "2026-04-03": "Day", "2026-04-04": "Off", "2026-04-05": "Day", "2026-04-06": "Day", "2026-04-07": "Day", "2026-04-08": "Day", "2026-04-09": "Day", "2026-04-10": "Day", "2026-04-11": "Off", "2026-04-12": "Day", "2026-04-13": "Day", "2026-04-14": "Day", "2026-04-15": "Day", "2026-04-16": "Day", "2026-04-17": "Day", "2026-04-18": "Off", "2026-04-19": "Day", "2026-04-20": "Day"}}, {"name": "Sajid", "role": "Technician", "schedule": {"2026-03-18": "Day", "2026-03-19": "Eid Leave", "2026-03-23": "Day", "2026-03-24": "Day", "2026-03-25": "Day", "2026-03-26": "Day", "2026-03-27": "Off", "2026-03-28": "Day", "2026-03-29": "Day", "2026-03-30": "Day", "2026-03-31": "Day", "2026-04-01": "Day", "2026-04-02": "Day", "2026-04-03": "Off", "2026-04-04": "Day", "2026-04-05": "Day", "2026-04-06": "Day", "2026-04-07": "Day", "2026-04-08": "Day", "2026-04-09": "Day", "2026-04-10": "Off", "2026-04-11": "Day", "2026-04-12": "Day", "2026-04-13": "Day", "2026-04-14": "Day", "2026-04-15": "Day", "2026-04-16": "Day", "2026-04-17": "Off", "2026-04-18": "Day", "2026-04-19": "Day", "2026-04-20": "Day"}}, {"name": "Haqnawaz", "role": "Technician", "schedule": {"2026-03-18": "Day", "2026-03-19": "Day", "2026-03-20": "Day", "2026-03-21": "Day", "2026-03-22": "Day", "2026-03-23": "Eid Leave", "2026-03-27": "Day", "2026-03-28": "Off", "2026-03-29": "Day", "2026-03-30": "Day", "2026-03-31": "Day", "2026-04-01": "Day", "2026-04-02": "Day", "2026-04-03": "Day", "2026-04-04": "Off", "2026-04-05": "Day", "2026-04-06": "Day", "2026-04-07": "Day", "2026-04-08": "Day", "2026-04-09": "Day", "2026-04-10": "Day", "2026-04-11": "Off", "2026-04-12": "Day", "2026-04-13": "Day", "2026-04-14": "Day", "2026-04-15": "Day", "2026-04-16": "Day", "2026-04-17": "Day", "2026-04-18": "Off", "2026-04-19": "Day", "2026-04-20": "Day"}}, {"name": "Arif", "role": "Technician", "schedule": {"2026-03-18": "Day", "2026-03-19": "Eid Leave", "2026-03-23": "Night", "2026-03-24": "Night", "2026-03-25": "Night", "2026-03-26": "Night", "2026-03-27": "Off", "2026-03-28": "Day", "2026-03-29": "Day", "2026-03-30": "Day", "2026-03-31": "Day", "2026-04-01": "Day", "2026-04-02": "Day", "2026-04-03": "Off", "2026-04-04": "Night", "2026-04-05": "Night", "2026-04-06": "Night", "2026-04-07": "Night", "2026-04-08": "Night", "2026-04-09": "Night", "2026-04-10": "Off", "2026-04-11": "Day", "2026-04-12": "Day", "2026-04-13": "Day", "2026-04-14": "Day", "2026-04-15": "Day", "2026-04-16": "Day", "2026-04-17": "Off", "2026-04-18": "Night", "2026-04-19": "Night", "2026-04-20": "Night"}}, {"name": "Majid", "role": "Technician", "schedule": {"2026-03-18": "Night", "2026-03-19": "Night", "2026-03-20": "Night", "2026-03-21": "Night", "2026-03-22": "Night", "2026-03-23": "Eid Leave", "2026-03-27": "Night", "2026-03-28": "Off", "2026-03-29": "Night", "2026-03-30": "Night", "2026-03-31": "Night", "2026-04-01": "Night", "2026-04-02": "Night", "2026-04-03": "Night", "2026-04-04": "Off", "2026-04-05": "Night", "2026-04-06": "Night", "2026-04-07": "Night", "2026-04-08": "Night", "2026-04-09": "Night", "2026-04-10": "Night", "2026-04-11": "Off", "2026-04-12": "Night", "2026-04-13": "Night", "2026-04-14": "Night", "2026-04-15": "Night", "2026-04-16": "Night", "2026-04-17": "Night", "2026-04-18": "Off", "2026-04-19": "Night", "2026-04-20": "Night"}}, {"name": "Shafiq", "role": "Technician", "schedule": {"2026-03-18": "Night", "2026-03-19": "Eid Leave", "2026-03-23": "Night", "2026-03-24": "Night", "2026-03-25": "Night", "2026-03-26": "Night", "2026-03-27": "Night", "2026-03-28": "Night", "2026-03-29": "Off", "2026-03-30": "Night", "2026-03-31": "Night", "2026-04-01": "Night", "2026-04-02": "Night", "2026-04-03": "Night", "2026-04-04": "Night", "2026-04-05": "Off", "2026-04-06": "Night", "2026-04-07": "Night", "2026-04-08": "Night", "2026-04-09": "Night", "2026-04-10": "Night", "2026-04-11": "Night", "2026-04-12": "Off", "2026-04-13": "Night", "2026-04-14": "Night", "2026-04-15": "Night", "2026-04-16": "Night", "2026-04-17": "Night", "2026-04-18": "Night", "2026-04-19": "Off", "2026-04-20": "Night"}}, {"name": "Mahiwal", "role": "Technician", "schedule": {"2026-03-18": "Night", "2026-03-19": "Night", "2026-03-20": "Night", "2026-03-21": "Night", "2026-03-22": "Night", "2026-03-23": "Eid Leave", "2026-03-27": "OFF", "2026-03-28": "Night", "2026-03-29": "Night", "2026-03-30": "Night", "2026-03-31": "Night", "2026-04-01": "Night", "2026-04-02": "Night", "2026-04-03": "OFF", "2026-04-04": "Night", "2026-04-05": "Night", "2026-04-06": "Night", "2026-04-07": "Night", "2026-04-08": "Night", "2026-04-09": "Night", "2026-04-10": "OFF", "2026-04-11": "Night", "2026-04-12": "Night", "2026-04-13": "Night", "2026-04-14": "Night", "2026-04-15": "Night", "2026-04-16": "Night", "2026-04-17": "OFF", "2026-04-18": "Night", "2026-04-19": "Night", "2026-04-20": "Night"}}, {"name": "Fazlulhaq", "role": "Technician", "schedule": {"2026-03-18": "Night", "2026-03-19": "Eid Leave", "2026-03-23": "Night", "2026-03-24": "Night", "2026-03-25": "Night", "2026-03-26": "Night", "2026-03-27": "Night", "2026-03-28": "Off", "2026-03-29": "Night", "2026-03-30": "Night", "2026-03-31": "Night", "2026-04-01": "Night", "2026-04-02": "Night", "2026-04-03": "Night", "2026-04-04": "Off", "2026-04-05": "Night", "2026-04-06": "Night", "2026-04-07": "Night", "2026-04-08": "Night", "2026-04-09": "Night", "2026-04-10": "Night", "2026-04-11": "Off", "2026-04-12": "Night", "2026-04-13": "Night", "2026-04-14": "Night", "2026-04-15": "Night", "2026-04-16": "Night", "2026-04-17": "Night", "2026-04-18": "Off", "2026-04-19": "Night", "2026-04-20": "Night"}}, {"name": "Muzammil", "role": "Technician", "schedule": {"2026-03-18": "Night", "2026-03-19": "Eid Leave", "2026-03-23": "Night", "2026-03-24": "Night", "2026-03-25": "Night", "2026-03-26": "Night", "2026-03-27": "Night", "2026-03-28": "Night", "2026-03-29": "Off", "2026-03-30": "Night", "2026-03-31": "Night", "2026-04-01": "Night", "2026-04-02": "Night", "2026-04-03": "Night", "2026-04-04": "Night", "2026-04-05": "Off", "2026-04-06": "Night", "2026-04-07": "Night", "2026-04-08": "Night", "2026-04-09": "Night", "2026-04-10": "Night", "2026-04-11": "Night", "2026-04-12": "Off", "2026-04-13": "Night", "2026-04-14": "Night", "2026-04-15": "Night", "2026-04-16": "Night", "2026-04-17": "Night", "2026-04-18": "Night", "2026-04-19": "Off", "2026-04-20": "Night"}}, {"name": "Rehan", "role": "Technician", "schedule": {"2026-03-18": "Night", "2026-03-19": "Eid Leave", "2026-03-23": "Night", "2026-03-24": "Night", "2026-03-25": "Night", "2026-03-26": "Night", "2026-03-27": "Night", "2026-03-28": "Off", "2026-03-29": "Night", "2026-03-30": "Night", "2026-03-31": "Night", "2026-04-01": "Night", "2026-04-02": "Night", "2026-04-03": "Night", "2026-04-04": "Off", "2026-04-05": "Night", "2026-04-06": "Night", "2026-04-07": "Night", "2026-04-08": "Night", "2026-04-09": "Night", "2026-04-10": "Night", "2026-04-11": "Off", "2026-04-12": "Night", "2026-04-13": "Night", "2026-04-14": "Night", "2026-04-15": "Night", "2026-04-16": "Night", "2026-04-17": "Night", "2026-04-18": "Off", "2026-04-19": "Night", "2026-04-20": "Night"}}, {"name": "Zain", "role": "Technician", "schedule": {"2026-03-18": "Night", "2026-03-19": "Eid Leave", "2026-03-23": "Night", "2026-03-24": "Night", "2026-03-25": "Night", "2026-03-26": "Night", "2026-03-27": "Night", "2026-03-28": "Off", "2026-03-29": "Night", "2026-03-30": "Night", "2026-03-31": "Night", "2026-04-01": "Night", "2026-04-02": "Night", "2026-04-03": "Night", "2026-04-04": "Off", "2026-04-05": "Night", "2026-04-06": "Night", "2026-04-07": "Night", "2026-04-08": "Night", "2026-04-09": "Night", "2026-04-10": "Night", "2026-04-11": "Off", "2026-04-12": "Night", "2026-04-13": "Night", "2026-04-14": "Night", "2026-04-15": "Night", "2026-04-16": "Night", "2026-04-17": "Night", "2026-04-18": "Off", "2026-04-19": "Night", "2026-04-20": "Night"}}, {"name": "Hamza", "role": "Technician", "schedule": {"2026-03-18": "Night", "2026-03-19": "Eid Leave", "2026-03-23": "Night", "2026-03-24": "Night", "2026-03-25": "Night", "2026-03-26": "Night", "2026-03-27": "OFF", "2026-03-28": "Night", "2026-03-29": "Off", "2026-03-30": "Night", "2026-03-31": "Night", "2026-04-01": "Night", "2026-04-02": "Night", "2026-04-03": "Night", "2026-04-04": "Night", "2026-04-05": "Off", "2026-04-06": "Night", "2026-04-07": "Night", "2026-04-08": "Night", "2026-04-09": "Night", "2026-04-10": "Night", "2026-04-11": "Night", "2026-04-12": "Off", "2026-04-13": "Night", "2026-04-14": "Night", "2026-04-15": "Night", "2026-04-16": "Night", "2026-04-17": "Night", "2026-04-18": "Night", "2026-04-19": "Off", "2026-04-20": "Night"}}, {"name": "Helper 1", "role": "Technician", "schedule": {"2026-03-18": "Night", "2026-03-19": "Eid Leave", "2026-03-23": "Night", "2026-03-24": "Night", "2026-03-25": "Night", "2026-03-26": "Night", "2026-03-27": "OFF", "2026-03-28": "Night", "2026-03-29": "Off", "2026-03-30": "Night", "2026-03-31": "Night", "2026-04-01": "Night", "2026-04-02": "Night", "2026-04-03": "Night", "2026-04-04": "Night", "2026-04-05": "Off", "2026-04-06": "Night", "2026-04-07": "Night", "2026-04-08": "Night", "2026-04-09": "Night", "2026-04-10": "Night", "2026-04-11": "Night", "2026-04-12": "Off", "2026-04-13": "Night", "2026-04-14": "Night", "2026-04-15": "Night", "2026-04-16": "Night", "2026-04-17": "Night", "2026-04-18": "Night", "2026-04-19": "Off", "2026-04-20": "Night"}}, {"name": "Helper 2", "role": "Technician", "schedule": {"2026-03-18": "Night", "2026-03-19": "Eid Leave", "2026-03-23": "Night", "2026-03-24": "Night", "2026-03-25": "Night", "2026-03-26": "Night", "2026-03-27": "OFF", "2026-03-28": "Night", "2026-03-29": "Off", "2026-03-30": "Night", "2026-03-31": "Night", "2026-04-01": "Night", "2026-04-02": "Night", "2026-04-03": "Night", "2026-04-04": "Night", "2026-04-05": "Off", "2026-04-06": "Night", "2026-04-07": "Night", "2026-04-08": "Night", "2026-04-09": "Night", "2026-04-10": "Night", "2026-04-11": "Night", "2026-04-12": "Off", "2026-04-13": "Night", "2026-04-14": "Night", "2026-04-15": "Night", "2026-04-16": "Night", "2026-04-17": "Night", "2026-04-18": "Night", "2026-04-19": "Off", "2026-04-20": "Night"}}, {"name": "Helper 3", "role": "Technician", "schedule": {"2026-03-18": "Night", "2026-03-19": "Eid Leave", "2026-03-23": "Night", "2026-03-24": "Night", "2026-03-25": "Night", "2026-03-26": "Night", "2026-03-27": "Night", "2026-03-28": "Off", "2026-03-29": "Night", "2026-03-30": "Night", "2026-03-31": "Night", "2026-04-01": "Night", "2026-04-02": "Night", "2026-04-03": "Night", "2026-04-04": "Off", "2026-04-05": "Night", "2026-04-06": "Night", "2026-04-07": "Night", "2026-04-08": "Night", "2026-04-09": "Night", "2026-04-10": "Night", "2026-04-11": "Off", "2026-04-12": "Night", "2026-04-13": "Night", "2026-04-14": "Night", "2026-04-15": "Night", "2026-04-16": "Night", "2026-04-17": "Night", "2026-04-18": "Off", "2026-04-19": "Night", "2026-04-20": "Night"}}, {"name": "Helper 4", "role": "Technician", "schedule": {"2026-03-18": "Night", "2026-03-19": "Eid Leave", "2026-03-23": "Night", "2026-03-24": "Night", "2026-03-25": "Night", "2026-03-26": "Night", "2026-03-27": "Night", "2026-03-28": "Night", "2026-03-29": "Off", "2026-03-30": "Night", "2026-03-31": "Night", "2026-04-01": "Night", "2026-04-02": "Night", "2026-04-03": "Night", "2026-04-04": "Night", "2026-04-05": "Off", "2026-04-06": "Night", "2026-04-07": "Night", "2026-04-08": "Night", "2026-04-09": "Night", "2026-04-10": "Night", "2026-04-11": "Night", "2026-04-12": "Off", "2026-04-13": "Night", "2026-04-14": "Night", "2026-04-15": "Night", "2026-04-16": "Night", "2026-04-17": "Night", "2026-04-18": "Night", "2026-04-19": "Off", "2026-04-20": "Night"}}]
        _default_eng_dates  = ["2026-03-19", "2026-03-20", "2026-03-21", "2026-03-22", "2026-03-23", "2026-03-24", "2026-03-25", "2026-03-26", "2026-03-27", "2026-03-28", "2026-03-29", "2026-03-30", "2026-03-31", "2026-04-01", "2026-04-02", "2026-04-03", "2026-04-04", "2026-04-05", "2026-04-06", "2026-04-07", "2026-04-08", "2026-04-09", "2026-04-10", "2026-04-11", "2026-04-12", "2026-04-13", "2026-04-14", "2026-04-15", "2026-04-16", "2026-04-17", "2026-04-18", "2026-04-19", "2026-04-20", "2026-04-21", "2026-04-22", "2026-04-23", "2026-04-24", "2026-04-25", "2026-04-26", "2026-04-27", "2026-04-28", "2026-04-29", "2026-04-30"]
        _default_tech_dates = ["2026-03-18", "2026-03-19", "2026-03-20", "2026-03-21", "2026-03-22", "2026-03-23", "2026-03-24", "2026-03-25", "2026-03-26", "2026-03-27", "2026-03-28", "2026-03-29", "2026-03-30", "2026-03-31", "2026-04-01", "2026-04-02", "2026-04-03", "2026-04-04", "2026-04-05", "2026-04-06", "2026-04-07", "2026-04-08", "2026-04-09", "2026-04-10", "2026-04-11", "2026-04-12", "2026-04-13", "2026-04-14", "2026-04-15", "2026-04-16", "2026-04-17", "2026-04-18", "2026-04-19", "2026-04-20"]
        ctx['engineers_json']   = _jd.dumps(_default_eng)
        ctx['technicians_json'] = _jd.dumps(_default_tech)
        ctx['eng_dates_json']   = _jd.dumps(_default_eng_dates)
        ctx['tech_dates_json']  = _jd.dumps(_default_tech_dates)
        ctx['data_source'] = 'default'
    return render(request, "core/manpower.html", ctx)

@login_required
def tracing_hub(request):
    redir = _require_module_page(request, 'tracing', 'view')
    if redir:
        return redir
    return render(request, "core/tracing_hub.html", _ctx(request, {"sheets": TRACING_SHEETS}))

@login_required
def tracing_sheet(request, slug):
    redir = _require_module_page(request, 'tracing', 'view')
    if redir:
        return redir
    sheet_info = SLUG_MAP.get(slug)
    if not sheet_info:
        from django.http import Http404
        raise Http404
    return render(request, "core/tracing_sheet.html",
                  _ctx(request, {"sheet": sheet_info, "sheets": TRACING_SHEETS}))

@login_required
def tracing_sheet_api(request, slug):
    _, err = _require_module_api(request, 'tracing', 'view')
    if err:
        return err
    sheet_info = SLUG_MAP.get(slug)
    if not sheet_info:
        return JsonResponse({"error": "Unknown sheet"}, status=404)
    raw, err = fetch_sheet_csv(sheet_info["sheet"])
    if err:
        return JsonResponse({"error": err}, status=500)
    data = parse_generic_sheet(raw)
    data.update(sheet_name=sheet_info["name"], color=sheet_info["color"])
    return JsonResponse(data)

@login_required
def documents(request):
    redir = _require_module_page(request, 'documents', 'view')
    if redir:
        return redir
    return render(request, "core/section.html",
                  _ctx(request, {"title": "Documents", "icon": "📄",
                                 "description": "Manage all your enterprise documents"}))

@login_required
def annual_plan(request):
    return render(request, "core/section.html",
                  _ctx(request, {"title": "Annual Plan", "icon": "📅",
                                 "description": "Strategic annual planning and goals"}))

@login_required
def daily_report(request):
    redir = _require_module_page(request, 'daily_report', 'view')
    if redir:
        return redir
    return render(request, "core/section.html",
                  _ctx(request, {"title": "Daily Report", "icon": "📊",
                                 "description": "Daily operational reports and insights"}))

@csrf_exempt
@login_required
def chat_api(request):
    if request.method == "POST":
        body = json.loads(request.body)
        message = body.get("message", "").lower()
        responses = {
            "manpower": "Man Power: 10 Engineers + 18 Technicians on the 1100MW Al Henakiya project.",
            "tracing":  "Maintenance Tracing has 11 live modules: PM, CM, SCB, Trackers, Inverter, Strings, PV, SS and more.",
            "document": "Document repository: 2,847 files. Latest uploads today at 09:15 AM.",
            "annual":   "HDEC Al Henakiya 2025 Annual Plan: 73% of Q1 complete.",
            "report":   "Today's report: 94.2% productivity, 18 tasks done.",
            "hello":    "Hello! I'm the HDEC Project Assistant for 1100MW Al Henakiya.",
            "help":     "I assist with: 👥 Manpower, 📡 Tracing, 📄 Documents, 📅 Annual Plan, 📊 Daily Reports.",
        }
        reply = "I'm the HDEC Project Assistant for 1100MW Al Henakiya. What would you like to explore?"
        for key, resp in responses.items():
            if key in message:
                reply = resp
                break
        return JsonResponse({"reply": reply})
    return JsonResponse({"error": "Method not allowed"}, status=405)

# ── ANNUAL PLAN ────────────────────────────────────────────────────────────

ANNUAL_PLAN_SHEET_ID = "1l418pkJkui6o_0Ib6r5-d0gr6AQrti_GocCr_GKxSwo"

# All tabs to fetch
ANNUAL_SHEETS = [
    {"slug": "pm-calendar",   "name": "PM Calendar",              "icon": "📅", "color": "#3b9eff",  "sheet": "PM Calander",                  "gid": "1182899491"},
    {"slug": "overall-pm",    "name": "Overall PM",               "icon": "📊", "color": "#00e5c8",  "sheet": "Overall PM",                   "gid": ""},
    {"slug": "pm-worksheet",  "name": "PM Worksheet",             "icon": "📋", "color": "#a259ff",  "sheet": "PM Worksheet",                 "gid": ""},
    {"slug": "daily-plan",    "name": "Daily Plan",               "icon": "📆", "color": "#f0c040",  "sheet": "Daily Plan",                   "gid": ""},
    {"slug": "checklist",     "name": "Checklist & Procedures",   "icon": "✅", "color": "#22c55e",  "sheet": "Checklist and Procedure Status","gid": ""},
    {"slug": "dec-25",        "name": "Dec-25 Report",            "icon": "📁", "color": "#fb923c",  "sheet": "Dec-25",                       "gid": ""},
]
ANNUAL_SLUG_MAP = {s["slug"]: s for s in ANNUAL_SHEETS}


def fetch_annual_csv(sheet_name):
    url = (f"https://docs.google.com/spreadsheets/d/{ANNUAL_PLAN_SHEET_ID}"
           f"/export?format=csv&sheet={urllib.parse.quote(sheet_name)}")
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as r:
            return r.read().decode("utf-8"), None
    except Exception as e:
        return None, str(e)


def parse_pm_calendar(raw_csv):
    """Parse the PM Calendar sheet into structured tasks grouped by equipment."""
    reader = csv.reader(io.StringIO(raw_csv))
    rows = list(reader)
    if len(rows) < 2:
        return {}

    # Row 0 = headers: S.No, Equipment, (blank), Task Description, Frequency, Date&Month, Start, End, [dates...]
    # Find actual data rows (row index 1 onward, skip blank rows)
    tasks = []
    for row in rows[1:]:
        if not row or not row[0].strip() or not row[0].strip().replace('.','').isdigit():
            continue
        sno       = row[0].strip()
        equipment = row[1].strip() if len(row) > 1 else ""
        task      = row[3].strip() if len(row) > 3 else ""
        frequency = row[4].strip() if len(row) > 4 else ""
        date_month= row[5].strip() if len(row) > 5 else ""
        start     = row[6].strip() if len(row) > 6 else ""
        end       = row[7].strip() if len(row) > 7 else ""
        if equipment and task:
            tasks.append({
                "sno": sno, "equipment": equipment, "task": task,
                "frequency": frequency, "date_month": date_month,
                "start": start, "end": end,
            })

    # Group by equipment
    groups = {}
    for t in tasks:
        eq = t["equipment"]
        if eq not in groups:
            groups[eq] = []
        groups[eq].append(t)

    return {
        "tasks": tasks,
        "groups": groups,
        "total": len(tasks),
        "equipment_list": list(groups.keys()),
        "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


@login_required
def annual_plan(request):
    return render(request, "core/annual_plan_hub.html",
                  _ctx(request, {"sheets": ANNUAL_SHEETS}))


@login_required
def annual_plan_sheet(request, slug):
    sheet_info = ANNUAL_SLUG_MAP.get(slug)
    if not sheet_info:
        from django.http import Http404
        raise Http404
    return render(request, "core/annual_plan_sheet.html",
                  _ctx(request, {"sheet": sheet_info, "sheets": ANNUAL_SHEETS}))


@login_required
def annual_plan_api(request, slug):
    sheet_info = ANNUAL_SLUG_MAP.get(slug)
    if not sheet_info:
        return JsonResponse({"error": "Unknown sheet"}, status=404)
    raw, err = fetch_annual_csv(sheet_info["sheet"])
    if err:
        return JsonResponse({"error": err}, status=500)
    if slug == "pm-calendar":
        data = parse_pm_calendar(raw)
    else:
        data = parse_generic_sheet(raw)
    data["sheet_name"] = sheet_info["name"]
    data["color"] = sheet_info["color"]
    data["fetched_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return JsonResponse(data)

# ── ANNUAL PLAN ─────────────────────────────────────────────────────────────

ANNUAL_SHEET_ID = "1l418pkJkui6o_0Ib6r5-d0gr6AQrti_GocCr_GKxSwo"

ANNUAL_SHEETS = [
    {"slug": "pm-calendar",    "name": "PM Calendar",                   "icon": "📅", "color": "#3b9eff",  "sheet": "PM Calander"},
    {"slug": "overall-pm",     "name": "Overall PM",                    "icon": "📊", "color": "#00e5c8",  "sheet": "Overall PM"},
    {"slug": "pm-worksheet",   "name": "PM Worksheet",                  "icon": "📋", "color": "#a259ff",  "sheet": "PM Worksheet"},
    {"slug": "daily-plan",     "name": "Daily Plan",                    "icon": "📆", "color": "#f0c040",  "sheet": "Daily Plan"},
    {"slug": "checklist",      "name": "Checklist & Procedure Status",  "icon": "✅", "color": "#22c55e",  "sheet": "Checklist and Procedure Status"},
]
ANNUAL_SLUG_MAP = {s["slug"]: s for s in ANNUAL_SHEETS}


def fetch_annual_csv(sheet_name):
    url = (f"https://docs.google.com/spreadsheets/d/{ANNUAL_SHEET_ID}"
           f"/export?format=csv&sheet={urllib.parse.quote(sheet_name)}")
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as r:
            return r.read().decode("utf-8"), None
    except Exception as e:
        return None, str(e)


def parse_pm_calendar(raw_csv):
    """Parse the PM Calendar sheet into structured task objects."""
    import csv as _csv, io as _io
    reader = _csv.reader(_io.StringIO(raw_csv))
    rows = list(reader)
    if len(rows) < 2:
        return []

    # Row 0 is headers: S.No, Equipment, (blank), Task Description, Frequency, Date&Month, Start, End, then dates...
    tasks = []
    for row in rows[1:]:
        if len(row) < 5 or not row[0].strip().isdigit():
            continue
        tasks.append({
            "sno":        row[0].strip(),
            "equipment":  row[1].strip(),
            "task":       row[3].strip() if len(row) > 3 else "",
            "frequency":  row[4].strip() if len(row) > 4 else "",
            "date_month": row[5].strip() if len(row) > 5 else "",
            "start_date": row[6].strip() if len(row) > 6 else "",
            "end_date":   row[7].strip() if len(row) > 7 else "",
        })
    return tasks


@login_required
def annual_plan(request):
    return render(request, "core/annual_hub.html",
                  _ctx(request, {"sheets": ANNUAL_SHEETS}))


@login_required
def annual_sheet(request, slug):
    sheet_info = ANNUAL_SLUG_MAP.get(slug)
    if not sheet_info:
        from django.http import Http404
        raise Http404
    return render(request, "core/annual_sheet.html",
                  _ctx(request, {"sheet": sheet_info, "sheets": ANNUAL_SHEETS}))


def annual_api(request, slug):
    sheet_info = ANNUAL_SLUG_MAP.get(slug)
    if not sheet_info:
        return JsonResponse({"error": "Unknown sheet"}, status=404)
    raw, err = fetch_annual_csv(sheet_info["sheet"])
    if err:
        return JsonResponse({"error": err}, status=500)

    if slug == "pm-calendar":
        tasks = parse_pm_calendar(raw)
        # Group by equipment
        from collections import defaultdict as _dd, Counter as _Cnt
        by_equipment = _dd(list)
        for t in tasks:
            by_equipment[t["equipment"]].append(t)
        freq_counts = dict(_Cnt(t["frequency"] for t in tasks).most_common())
        equip_counts = {k: len(v) for k, v in by_equipment.items()}
        # Month distribution from date_month field
        month_counts = dict(_Cnt(t["date_month"] for t in tasks if t["date_month"]).most_common())
        return JsonResponse({
            "tasks": tasks,
            "by_equipment": {k: v for k, v in by_equipment.items()},
            "freq_counts": freq_counts,
            "equip_counts": equip_counts,
            "month_counts": month_counts,
            "total": len(tasks),
            "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
    else:
        data = parse_generic_sheet(raw)
        data["sheet_name"] = sheet_info["name"]
        return JsonResponse(data)

# ═══════════════════════════════════════════════════════════════════════════
# ANNUAL PLAN
# ═══════════════════════════════════════════════════════════════════════════

from .annual_plan_data import (
    ANNUAL_SHEET_ID, ANNUAL_SHEETS, EQUIPMENT_FOLDERS,
    EQUIP_SLUG_MAP, SHEET_SLUG_MAP, FREQ_COLORS
)

def _fetch_pm_calendar():
    """Fetch PM Calendar sheet and parse all 73 tasks."""
    url = (f"https://docs.google.com/spreadsheets/d/{ANNUAL_SHEET_ID}"
           f"/export?format=csv&sheet=PM%20Calander")
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as r:
            raw = r.read().decode("utf-8")
    except Exception as e:
        return None, str(e)

    reader = csv.reader(io.StringIO(raw))
    rows = list(reader)
    if not rows:
        return None, "Empty response"

    # Row 0 is header (S.No, Equipment, _, Task Description, Frequency, Date&Month, Start, End, date1, date2...)
    header = rows[0]
    # Date columns start at index 8
    date_cols = []
    for h in header[8:]:
        h = h.strip()
        if h:
            date_cols.append(h)

    tasks = []
    for row in rows[1:]:
        if not row or not row[0].strip() or not row[0].strip().isdigit():
            continue
        sno       = row[0].strip()
        equipment = row[1].strip() if len(row) > 1 else ""
        task_desc = row[3].strip() if len(row) > 3 else ""
        frequency = row[4].strip() if len(row) > 4 else ""
        date_month= row[5].strip() if len(row) > 5 else ""
        start_date= row[6].strip() if len(row) > 6 else ""
        end_date  = row[7].strip() if len(row) > 7 else ""

        # Daily status values (P=done, empty=pending etc)
        daily = {}
        for i, dc in enumerate(date_cols):
            val = row[8+i].strip() if (8+i) < len(row) else ""
            if val:
                daily[dc] = val

        tasks.append({
            "sno": sno, "equipment": equipment, "task": task_desc,
            "frequency": frequency, "date_month": date_month,
            "start": start_date, "end": end_date, "daily": daily,
        })

    return {"tasks": tasks, "date_cols": date_cols,
            "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}, None


@login_required
def annual_plan(request):
    redir = _require_module_page(request, 'annual_plan', 'view')
    if redir:
        return redir
    return render(request, "core/annual_plan_hub.html",
                  _ctx(request, {"folders": EQUIPMENT_FOLDERS, "sheets": ANNUAL_SHEETS}))


@login_required
def annual_plan_folder(request, slug):
    redir = _require_module_page(request, 'annual_plan', 'view')
    if redir:
        return redir
    folder = EQUIP_SLUG_MAP.get(slug)
    if not folder:
        from django.http import Http404
        raise Http404
    return render(request, "core/annual_plan_folder.html",
                  _ctx(request, {"folder": folder, "folders": EQUIPMENT_FOLDERS}))


@login_required
def annual_plan_sheet(request, slug):
    redir = _require_module_page(request, 'annual_plan', 'view')
    if redir:
        return redir
    sheet = SHEET_SLUG_MAP.get(slug)
    if not sheet:
        from django.http import Http404
        raise Http404
    return render(request, "core/annual_plan_sheet.html",
                  _ctx(request, {"sheet": sheet, "sheets": ANNUAL_SHEETS, "folders": EQUIPMENT_FOLDERS}))


def annual_plan_api(request):
    """Full PM Calendar JSON — used by hub overview."""
    _, err = _require_module_api(request, 'annual_plan', 'view')
    if err:
        return err
    data, err = _fetch_pm_calendar()
    if err:
        return JsonResponse({"error": err}, status=500)
    return JsonResponse(data)


def annual_plan_folder_api(request, slug):
    """Tasks for a specific equipment folder."""
    _, err = _require_module_api(request, 'annual_plan', 'view')
    if err:
        return err
    folder = EQUIP_SLUG_MAP.get(slug)
    if not folder:
        return JsonResponse({"error": "Not found"}, status=404)
    data, err = _fetch_pm_calendar()
    if err:
        return JsonResponse({"error": err}, status=500)
    tasks = [t for t in data["tasks"] if t["equipment"] == folder["equip"]]
    freq_counts = dict(Counter(t["frequency"] for t in tasks))
    return JsonResponse({
        "folder": folder["name"], "equip": folder["equip"],
        "color": folder["color"], "tasks": tasks,
        "date_cols": data["date_cols"],
        "freq_counts": freq_counts,
        "total": len(tasks),
        "fetched_at": data["fetched_at"],
    })


def annual_plan_sheet_api(request, slug):
    """Fetch an individual named sheet."""
    _, err = _require_module_api(request, 'annual_plan', 'view')
    if err:
        return err
    sheet = SHEET_SLUG_MAP.get(slug)
    if not sheet:
        return JsonResponse({"error": "Not found"}, status=404)
    url = (f"https://docs.google.com/spreadsheets/d/{ANNUAL_SHEET_ID}"
           f"/export?format=csv&sheet={urllib.parse.quote(sheet['sheet'])}")
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as r:
            raw = r.read().decode("utf-8")
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    parsed = parse_generic_sheet(raw)
    parsed["sheet_name"] = sheet["name"]
    return JsonResponse(parsed)

# ═══════════════════════════════════════════════════════════════════════════
# MANPOWER — IMPORT / EXPORT EXCEL
# ═══════════════════════════════════════════════════════════════════════════

import io as _io
import json as _json
from datetime import datetime as _dt
from pathlib import Path as _Path

SCHEDULE_FILE = _Path(__file__).resolve().parent.parent / 'schedule_store.json'

def _load_schedule():
    if SCHEDULE_FILE.exists():
        try:
            with open(SCHEDULE_FILE) as f:
                return _json.load(f)
        except Exception:
            pass
    return None

def _save_schedule(data):
    with open(SCHEDULE_FILE, 'w') as f:
        _json.dump(data, f)

def _api_auth(request, admin_only=False, module_id=None, level='view'):
    """Check session for API endpoints. Returns user dict or None."""
    user = get_user(request)
    if not user:
        return None, JsonResponse({'error': 'Not authenticated'}, status=401)
    if module_id:
        if not _has_any_project_access(user) or not has_permission(user, module_id, level):
            return None, JsonResponse({'error': 'Forbidden'}, status=403)
    if admin_only and user.get('role') != 'admin':
        return None, JsonResponse({'error': 'Admin access required'}, status=403)
    return user, None

@csrf_exempt
def manpower_import(request):
    """Accept uploaded .xlsx and parse into schedule JSON."""
    _, err = _api_auth(request, admin_only=True, module_id='manpower', level='edit')
    if err:
        return err
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)

    f = request.FILES.get('file')
    if not f:
        return JsonResponse({'error': 'No file uploaded'}, status=400)
    if not f.name.endswith(('.xlsx', '.xlsm', '.xls')):
        return JsonResponse({'error': 'Must be an Excel file (.xlsx)'}, status=400)

    try:
        from openpyxl import load_workbook
    except ImportError:
        return JsonResponse({'error': 'openpyxl not installed. Run: pip install openpyxl'}, status=500)

    try:
        wb = load_workbook(_io.BytesIO(f.read()), read_only=True)
    except Exception as e:
        return JsonResponse({'error': f'Cannot read file: {e}'}, status=400)

    engineers = []
    technicians = []

    # ── Parse Engineer sheet ──
    eng_sheet_name = None
    for name in wb.sheetnames:
        if 'engineer' in name.lower():
            eng_sheet_name = name
            break
    if eng_sheet_name:
        ws = wb[eng_sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = rows[0]
            dates = []
            for c in header[3:]:
                if c and hasattr(c, 'strftime'):
                    dates.append(c.strftime('%Y-%m-%d'))
                elif c and isinstance(c, str) and c.strip():
                    dates.append(c.strip())
            for row in rows[1:]:
                if not row[2]:
                    continue
                schedule = {}
                for i, d in enumerate(dates):
                    val = row[3+i] if (3+i) < len(row) else None
                    if val and isinstance(val, str) and val.strip():
                        schedule[d] = val.strip()
                engineers.append({
                    'dept': str(row[0] or ''), 'role': str(row[1] or ''),
                    'name': str(row[2]), 'schedule': schedule
                })

    # ── Parse Technician sheet ──
    tech_sheet_name = None
    for name in wb.sheetnames:
        if 'tech' in name.lower():
            tech_sheet_name = name
            break
    if tech_sheet_name:
        ws2 = wb[tech_sheet_name]
        rows2 = list(ws2.iter_rows(values_only=True))
        if len(rows2) > 1:
            # Read ALL columns (no hardcoded cap) — skip non-name header values
            SKIP_WORDS = {'available', 'night', 'day', 'off', 'shift', 'name', 'date', 'total', ''}
            tech_names = []
            header_row = rows2[1] if len(rows2) > 1 else []
            for n in header_row[1:]:  # skip first column (date/label column)
                if n and isinstance(n, str) and n.strip().lower() not in SKIP_WORDS:
                    tech_names.append(n.strip())
            tech_schedule = {n: {} for n in tech_names}
            for row in rows2[2:]:
                if not row[0]:
                    continue
                # Accept both datetime objects and date strings
                if hasattr(row[0], 'strftime'):
                    d = row[0].strftime('%Y-%m-%d')
                elif isinstance(row[0], str) and row[0].strip():
                    d = row[0].strip()[:10]
                else:
                    continue
                for i, name in enumerate(tech_names):
                    col_idx = 1 + i
                    val = row[col_idx] if col_idx < len(row) else None
                    if val is None:
                        continue
                    val_str = str(val).strip()
                    if val_str and not val_str.startswith('='):
                        tech_schedule[name][d] = val_str
            technicians = [{'name': n, 'role': 'Technician', 'schedule': tech_schedule[n]} for n in tech_names]

    if not engineers and not technicians:
        return JsonResponse({'error': 'No data found. Expected sheets named "Engineer" and "Technician".'}, status=400)

    data = {
        'engineers': engineers, 'technicians': technicians,
        'imported_at': _dt.now().strftime('%Y-%m-%d %H:%M:%S'),
        'source_file': f.name,
    }
    _save_schedule(data)
    return JsonResponse({
        'ok': True,
        'engineers': len(engineers),
        'technicians': len(technicians),
        'imported_at': data['imported_at'],
        'source_file': f.name,
    })


@csrf_exempt
def manpower_export(request):
    """Export current schedule as formatted .xlsx. Accepts GET or POST."""
    _, err = _api_auth(request, module_id='manpower', level='view')
    if err:
        return err
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse('openpyxl not installed. Run: pip install openpyxl', status=500)

    # Load saved schedule or fall back to embedded data
    stored = _load_schedule()

    # POST = JS sends current in-memory data; GET = use stored data
    body_data = None
    if request.method == 'POST' and request.body:
        try:
            body_data = _json.loads(request.body)
        except Exception:
            pass

    if body_data and (body_data.get('engineers') or body_data.get('technicians')):
        engineers   = body_data.get('engineers', [])
        technicians = body_data.get('technicians', [])
    elif stored:
        engineers   = stored.get('engineers', [])
        technicians = stored.get('technicians', [])
    else:
        engineers   = []
        technicians = []

    wb = Workbook()

    # ── Color map ──
    SHIFT_FILLS = {
        'Day':       PatternFill('solid', fgColor='1A6BFF'),
        'Night':     PatternFill('solid', fgColor='7C3AED'),
        'Rest':      PatternFill('solid', fgColor='374151'),
        'General':   PatternFill('solid', fgColor='047857'),
        'OFF':       PatternFill('solid', fgColor='991B1B'),
        'ON leave':  PatternFill('solid', fgColor='92400E'),
        'EID leave': PatternFill('solid', fgColor='B45309'),
        'Eid Leave': PatternFill('solid', fgColor='B45309'),
    }
    WHITE  = Font(color='FFFFFF', bold=False, size=9)
    BOLD_W = Font(color='FFFFFF', bold=True, size=10)
    HDR_FILL = PatternFill('solid', fgColor='0F172A')
    HDR_FONT = Font(color='94A3B8', bold=True, size=9)
    thin = Side(style='thin', color='1E293B')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=False)
    left   = Alignment(horizontal='left',   vertical='center')

    def style_header(cell, text):
        cell.value = text
        cell.font  = HDR_FONT
        cell.fill  = HDR_FILL
        cell.alignment = center
        cell.border = bdr

    def write_sheet(ws, people, label):
        ws.title = label
        if not people:
            ws['A1'] = 'No data'
            return

        # Collect all dates across everyone
        all_dates = sorted(set(
            d for p in people for d in p.get('schedule', {})
        ))

        # Header row
        fixed_cols = ['Dept/Type', 'Role', 'Name'] if 'role' in people[0] and people[0]['role'] != 'Technician' else ['Name', 'Role']
        headers = fixed_cols + all_dates
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci)
            # Shorten date headers: "21-Nov" style
            if '-' in h and h.count('-') == 2:
                try:
                    dt = _dt.strptime(h, '%Y-%m-%d')
                    display = f"{dt.day}-{dt.strftime('%b')}"
                except Exception:
                    display = h
            else:
                display = h
            style_header(cell, display)

        # Freeze header
        ws.freeze_panes = 'D2' if len(fixed_cols) == 3 else 'C2'

        # Data rows
        for ri, person in enumerate(people, 2):
            sched = person.get('schedule', {})
            if len(fixed_cols) == 3:
                ws.cell(ri, 1, person.get('dept', '')).alignment = left
                ws.cell(ri, 2, person.get('role', '')).alignment = left
                ws.cell(ri, 3, person.get('name', '')).font = Font(bold=True, color='E2E8F5', size=10)
                ws.cell(ri, 3).alignment = left
                date_start = 4
            else:
                ws.cell(ri, 1, person.get('name', '')).font = Font(bold=True, color='E2E8F5', size=10)
                ws.cell(ri, 1).alignment = left
                ws.cell(ri, 2, person.get('role', 'Technician')).alignment = left
                date_start = 3

            for di, date in enumerate(all_dates):
                val = sched.get(date, '')
                cell = ws.cell(ri, date_start + di, val)
                cell.alignment = center
                cell.border = bdr
                cell.font = WHITE
                fill = SHIFT_FILLS.get(val)
                if fill:
                    cell.fill = fill
                else:
                    cell.fill = PatternFill('solid', fgColor='0F172A')

            # Row background for non-date cells
            bg = PatternFill('solid', fgColor='0D1424')
            for ci in range(1, date_start):
                c = ws.cell(ri, ci)
                c.fill = bg
                c.border = bdr

        # Column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 26
        if len(fixed_cols) == 3:
            ws.column_dimensions['C'].width = 18
        for di in range(len(all_dates)):
            col_letter = get_column_letter(date_start + di)
            ws.column_dimensions[col_letter].width = 7
        ws.row_dimensions[1].height = 28

        # Tab color
        ws.sheet_properties.tabColor = '3B9EFF'

    # Sheet 1 — Engineers
    ws_eng = wb.active
    write_sheet(ws_eng, engineers, 'Maintenance Engineers')

    # Sheet 2 — Technicians
    ws_tech = wb.create_sheet('Technicians')
    write_sheet(ws_tech, technicians, 'Technicians')

    # Sheet 3 — Summary
    ws_sum = wb.create_sheet('Summary')
    ws_sum.title = 'Summary'
    ws_sum.sheet_properties.tabColor = '00E5C8'
    summary_data = [
        ('HDEC — 1100MW Al Henakiya Project', ''),
        ('Man Power Schedule Export', ''),
        ('', ''),
        ('Exported At', _dt.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('Total Engineers', len(engineers)),
        ('Total Technicians', len(technicians)),
        ('Total Staff', len(engineers) + len(technicians)),
    ]
    for ri, (k, v) in enumerate(summary_data, 1):
        ws_sum.cell(ri, 1, k).font  = Font(color='94A3B8', bold=True, size=11)
        ws_sum.cell(ri, 2, v).font  = Font(color='E2E8F5', size=11)
        ws_sum.cell(ri, 1).fill = PatternFill('solid', fgColor='0F172A')
        ws_sum.cell(ri, 2).fill = PatternFill('solid', fgColor='0F172A')
    ws_sum.column_dimensions['A'].width = 28
    ws_sum.column_dimensions['B'].width = 28

    # Stream response
    from django.http import HttpResponse
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'HDEC_ManPower_{_dt.now().strftime("%Y%m%d_%H%M")}.xlsx'
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp

# ═══════════════════════════════════════════════════════════════════════════
# ATTENDANCE SYSTEM
# ═══════════════════════════════════════════════════════════════════════════

FACES_FILE      = _Path(__file__).resolve().parent.parent / 'face_descriptors.json'
ATTENDANCE_FILE = _Path(__file__).resolve().parent.parent / 'attendance_records.json'

def _load_faces():
    if FACES_FILE.exists():
        try:
            with open(FACES_FILE) as f:
                return _json.load(f)
        except Exception:
            pass
    return {}

def _save_faces(data):
    with open(FACES_FILE, 'w') as f:
        _json.dump(data, f)

def _load_attendance():
    if ATTENDANCE_FILE.exists():
        try:
            with open(ATTENDANCE_FILE) as f:
                return _json.load(f)
        except Exception:
            pass
    return {}

def _save_attendance(data):
    with open(ATTENDANCE_FILE, 'w') as f:
        _json.dump(data, f)


@csrf_exempt
def attendance_face_api(request):
    """GET: list trained faces. POST: save a new face descriptor."""
    _, err = _api_auth(request)
    if err:
        return err

    if request.method == 'GET':
        faces = _load_faces()
        return JsonResponse({'faces': list(faces.keys()), 'count': len(faces)})

    if request.method == 'POST':
        _, err = _api_auth(request, admin_only=True)
        if err:
            return err
        data = _json.loads(request.body)
        name       = data.get('name', '').strip()
        descriptor = data.get('descriptor', [])   # 128-float array
        label      = data.get('label', name)
        if not name or not descriptor:
            return JsonResponse({'error': 'name and descriptor required'}, status=400)
        if len(descriptor) != 128:
            return JsonResponse({'error': f'descriptor must be 128 floats, got {len(descriptor)}'}, status=400)
        faces = _load_faces()
        # Store multiple descriptors per person (up to 5 samples)
        if name not in faces:
            faces[name] = {'label': label, 'descriptors': []}
        if len(faces[name]['descriptors']) < 10:
            faces[name]['descriptors'].append(descriptor)
        else:
            # Replace oldest
            faces[name]['descriptors'].pop(0)
            faces[name]['descriptors'].append(descriptor)
        _save_faces(faces)
        return JsonResponse({'ok': True, 'name': name,
                             'samples': len(faces[name]['descriptors'])})

    return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
def attendance_face_delete(request):
    """DELETE a person's face data."""
    _, err = _api_auth(request, admin_only=True)
    if err:
        return err
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    data = _json.loads(request.body)
    name = data.get('name', '').strip()
    faces = _load_faces()
    if name in faces:
        del faces[name]
        _save_faces(faces)
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'Not found'}, status=404)


@csrf_exempt
def attendance_mark(request):
    """Mark time-in or time-out for a person."""
    _, err = _api_auth(request)
    if err:
        return err
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)

    data      = _json.loads(request.body)
    name      = data.get('name', '').strip()
    action    = data.get('action', 'in')    # 'in' or 'out'
    time_str  = data.get('time', _dt.now().strftime('%H:%M:%S'))
    date_str  = data.get('date', _dt.now().strftime('%Y-%m-%d'))

    # Location fields (sent from browser Geolocation API)
    lat           = data.get('lat')         # float or None
    lng           = data.get('lng')         # float or None
    location_name = data.get('location_name', '')   # reverse-geocoded name (if browser got it)
    location_accuracy = data.get('accuracy')        # metres

    if not name:
        return JsonResponse({'error': 'name required'}, status=400)

    records = _load_attendance()
    if date_str not in records:
        records[date_str] = {}

    if name not in records[date_str]:
        records[date_str][name] = {'time_in': None, 'time_out': None, 'status': 'Absent'}

    rec = records[date_str][name]

    # Build location snapshot for this action
    loc_snapshot = {}
    if lat is not None and lng is not None:
        loc_snapshot = {
            'lat': round(float(lat), 6),
            'lng': round(float(lng), 6),
            'accuracy': round(float(location_accuracy), 1) if location_accuracy else None,
            'name': location_name or f"{round(float(lat),5)}, {round(float(lng),5)}",
        }

    if action == 'in':
        rec['time_in']       = time_str
        rec['status']        = 'Present'
        if loc_snapshot:
            rec['location_in'] = loc_snapshot
    elif action == 'out':
        rec['time_out'] = time_str
        if loc_snapshot:
            rec['location_out'] = loc_snapshot
        if rec.get('time_in'):
            # Calculate hours
            try:
                fmt = '%H:%M:%S'
                ti  = _dt.strptime(rec['time_in'], fmt)
                to  = _dt.strptime(time_str, fmt)
                diff = to - ti
                hours = round(diff.seconds / 3600, 2)
                rec['hours'] = hours
            except Exception:
                pass

    _save_attendance(records)
    return JsonResponse({'ok': True, 'record': rec, 'date': date_str, 'name': name})


def attendance_get(request):
    """GET attendance for a date or month."""
    _, err = _api_auth(request)
    if err:
        return err

    date  = request.GET.get('date', _dt.now().strftime('%Y-%m-%d'))
    month = request.GET.get('month', '')   # YYYY-MM

    records = _load_attendance()

    if month:
        # Return all records for the month
        month_records = {d: v for d, v in records.items() if d.startswith(month)}
        return JsonResponse({'records': month_records, 'month': month})

    day_records = records.get(date, {})
    return JsonResponse({'records': day_records, 'date': date,
                         'fetched_at': _dt.now().strftime('%Y-%m-%d %H:%M:%S')})


def attendance_export(request):
    """Export attendance to Excel — daily or monthly."""
    _, err = _api_auth(request)
    if err:
        return err

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse('openpyxl not installed', status=500)

    date  = request.GET.get('date', _dt.now().strftime('%Y-%m-%d'))
    month = request.GET.get('month', '')
    mode  = 'monthly' if month else 'daily'

    records = _load_attendance()

    wb = Workbook()
    ws = wb.active

    # Styles
    HDR_FONT = Font(color='E2E8F5', bold=True, size=10)
    HDR_FILL = PatternFill('solid', fgColor='0F172A')
    PRES_FILL= PatternFill('solid', fgColor='14532D')
    ABS_FILL = PatternFill('solid', fgColor='450A0A')
    thin = Side(style='thin', color='1E293B')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center')
    lft  = Alignment(horizontal='left',   vertical='center')

    def hdr(cell, val, width=None):
        cell.value = val; cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = ctr; cell.border = bdr
        if width:
            ws.column_dimensions[get_column_letter(cell.column)].width = width

    if mode == 'daily':
        ws.title = f'Attendance {date}'
        day_recs = records.get(date, {})

        hdr(ws['A1'], 'S.No',          6)
        hdr(ws['B1'], 'Name',          22)
        hdr(ws['C1'], 'Date',          14)
        hdr(ws['D1'], 'Status',        12)
        hdr(ws['E1'], 'Time In',       12)
        hdr(ws['F1'], 'Time Out',      12)
        hdr(ws['G1'], 'Hours',         10)
        hdr(ws['H1'], 'Location In',   30)
        hdr(ws['I1'], 'Coords In',     22)
        hdr(ws['J1'], 'Location Out',  30)
        hdr(ws['K1'], 'Coords Out',    22)

        # Get all technician names
        # Use people roster (attendance_people.json) — falls back to schedule
        people_roster = _load_people()
        tech_names = [p['name'] for p in people_roster] if people_roster else sorted(day_recs.keys())
        if not tech_names:
            tech_names = sorted(day_recs.keys())

        for i, name in enumerate(tech_names, 1):
            rec = day_recs.get(name, {})
            status   = rec.get('status', 'Absent')
            time_in  = rec.get('time_in', '—')
            time_out = rec.get('time_out', '—')
            hours    = rec.get('hours', '—')
            loc_in   = rec.get('location_in', {})
            loc_out  = rec.get('location_out', {})
            loc_in_name  = loc_in.get('name', '') if loc_in else ''
            loc_in_coord = f"{loc_in['lat']}, {loc_in['lng']}" if loc_in else ''
            loc_out_name = loc_out.get('name', '') if loc_out else ''
            loc_out_coord= f"{loc_out['lat']}, {loc_out['lng']}" if loc_out else ''
            fill = PRES_FILL if status == 'Present' else ABS_FILL
            row_data = [i, name, date, status, time_in, time_out, hours,
                        loc_in_name, loc_in_coord, loc_out_name, loc_out_coord]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(i+1, ci, val)
                cell.fill = fill; cell.border = bdr
                cell.alignment = ctr if ci != 2 else lft
                cell.font = Font(color='E2E8F5', size=10)

        ws.row_dimensions[1].height = 26
        filename = f'HDEC_Attendance_{date}.xlsx'

    else:  # monthly
        # Parse year/month and generate ALL days in the month
        try:
            yr, mo = int(month[:4]), int(month[5:7])
        except Exception:
            yr, mo = _dt.now().year, _dt.now().month
        import calendar as _cal
        days_in_month = _cal.monthrange(yr, mo)[1]
        # Cap at today if exporting current month
        today = _dt.now()
        if yr == today.year and mo == today.month:
            days_in_month = today.day
        all_dates = [f'{yr:04d}-{mo:02d}-{d:02d}' for d in range(1, days_in_month + 1)]

        month_recs = {d: v for d, v in records.items() if d.startswith(month)}

        ws.title = f'Att {month}'

        people_roster = _load_people()
        tech_names = [p['name'] for p in people_roster] if people_roster else []
        if not tech_names:
            names_set = set()
            for dr in month_recs.values():
                names_set.update(dr.keys())
            tech_names = sorted(names_set)
        if not tech_names:
            tech_names = ['No people in roster']

        UNK_FILL  = PatternFill('solid', fgColor='1E293B')
        NAME_FILL = PatternFill('solid', fgColor='0D1424')
        WHITE_FONT     = Font(color='E2E8F5', size=10, bold=True)
        SMALL_FONT     = Font(color='E2E8F5', size=9)
        SUMMARY_FONT   = Font(color='E2E8F5', size=10)

        # Row 1 — title banner
        total_cols = 1 + len(all_dates) + 3  # Name + days + Present/Absent/Hours
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(1, 1, f'HDEC Attendance — {_dt.strptime(month, "%Y-%m").strftime("%B %Y")}')
        title_cell.font  = Font(color='E2E8F5', bold=True, size=12)
        title_cell.fill  = PatternFill('solid', fgColor='060D1A')
        title_cell.alignment = ctr
        ws.row_dimensions[1].height = 28

        # Row 2 — column headers
        hdr(ws.cell(2, 1), 'Name', 22)
        for di, d in enumerate(all_dates, 2):
            dt_obj = _dt.strptime(d, '%Y-%m-%d')
            day_label = f"{dt_obj.day}\n{dt_obj.strftime('%a')[:2]}"
            c = ws.cell(2, di, day_label)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); c.border = bdr
            ws.column_dimensions[get_column_letter(di)].width = 5.5
        sum_col = len(all_dates) + 2
        for offset, label in enumerate(['Present', 'Absent', 'Hours']):
            c = ws.cell(2, sum_col + offset, label)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = ctr; c.border = bdr
            ws.column_dimensions[get_column_letter(sum_col + offset)].width = 9
        ws.row_dimensions[2].height = 30

        # Data rows
        for ri, name in enumerate(tech_names, 3):
            c = ws.cell(ri, 1, name)
            c.font = WHITE_FONT; c.fill = NAME_FILL; c.alignment = lft; c.border = bdr

            present = 0; absent = 0; total_hours = 0.0
            for di, d in enumerate(all_dates, 2):
                day_data = month_recs.get(d)
                if day_data is None:
                    # Whole day has no records — Unknown
                    display = '?'; fill = UNK_FILL
                    is_present = False
                else:
                    rec = day_data.get(name)
                    if rec is None:
                        display = 'A'; fill = ABS_FILL; is_present = False
                    else:
                        ti  = rec.get('time_in', '')
                        hrs = rec.get('hours', 0)
                        is_present = bool(ti)
                        display = ti[:5] if ti else 'A'
                        fill    = PRES_FILL if is_present else ABS_FILL
                        try: total_hours += float(hrs)
                        except: pass

                if is_present: present += 1
                elif day_data is not None: absent += 1  # don't count unknown days as absent

                c = ws.cell(ri, di, display)
                c.fill = fill; c.alignment = ctr; c.border = bdr; c.font = SMALL_FONT

            for offset, val in enumerate([present, absent, round(total_hours, 1)]):
                c = ws.cell(ri, sum_col + offset, val)
                c.fill = NAME_FILL; c.alignment = ctr; c.border = bdr; c.font = SUMMARY_FONT

        # Freeze panes at B3
        ws.freeze_panes = 'B3'
        ws.row_dimensions[2].height = 30
        filename = f'HDEC_Attendance_{month}.xlsx'

    from django.http import HttpResponse
    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp

def attendance_face_descriptors(request):
    """Return full face descriptors (needed for FaceMatcher in browser)."""
    _, err = _api_auth(request)
    if err:
        return err
    return JsonResponse(_load_faces())

# ── ATTENDANCE PEOPLE MANAGEMENT ──────────────────────────────────────────

PEOPLE_FILE = _Path(__file__).resolve().parent.parent / 'attendance_people.json'

def _load_people():
    """Load attendance roster. Falls back to technicians from schedule."""
    if PEOPLE_FILE.exists():
        try:
            with open(PEOPLE_FILE) as f:
                return _json.load(f)
        except Exception:
            pass
    # Derive from schedule if exists
    stored = _load_schedule()
    if stored:
        return [{'name': p['name'], 'role': p.get('role','Technician')}
                for p in stored.get('technicians', [])]
    return []

def _save_people(data):
    with open(PEOPLE_FILE, 'w') as f:
        _json.dump(data, f, indent=2)


@csrf_exempt
def attendance_people(request):
    """GET: list people. POST: add. DELETE via POST with action=delete."""
    _, err = _api_auth(request)
    if err:
        return err

    if request.method == 'GET':
        return JsonResponse({'people': _load_people()})

    _, err = _api_auth(request, admin_only=True)
    if err:
        return err

    if request.method == 'POST':
        data   = _json.loads(request.body)
        action = data.get('action', 'add')

        if action == 'add':
            name = data.get('name', '').strip()
            role = data.get('role', 'Technician').strip()
            if not name:
                return JsonResponse({'error': 'Name is required'}, status=400)
            people = _load_people()
            if any(p['name'].lower() == name.lower() for p in people):
                return JsonResponse({'error': f'"{name}" already exists'}, status=400)
            people.append({'name': name, 'role': role})
            _save_people(people)
            return JsonResponse({'ok': True, 'people': people})

        elif action == 'delete':
            name = data.get('name', '').strip()
            people = _load_people()
            people = [p for p in people if p['name'] != name]
            _save_people(people)
            return JsonResponse({'ok': True, 'people': people})

        elif action == 'sync_from_schedule':
            # Re-sync from imported schedule
            stored = _load_schedule()
            if not stored:
                return JsonResponse({'error': 'No schedule imported yet'}, status=400)
            people = [{'name': p['name'], 'role': p.get('role','Technician')}
                      for p in stored.get('technicians', [])]
            _save_people(people)
            return JsonResponse({'ok': True, 'people': people, 'count': len(people)})

    return JsonResponse({'error': 'Method not allowed'}, status=405)

# ── FACE PHOTOS ───────────────────────────────────────────────────────────
PHOTOS_FILE = _Path(__file__).resolve().parent.parent / 'face_photos.json'

def _load_photos():
    if PHOTOS_FILE.exists():
        try:
            with open(PHOTOS_FILE) as f:
                return _json.load(f)
        except Exception:
            pass
    return {}

def _save_photos(data):
    with open(PHOTOS_FILE, 'w') as f:
        _json.dump(data, f)


@csrf_exempt
def attendance_face_photo_save(request):
    """Save a face photo (base64 JPEG) for a person during training."""
    _, err = _api_auth(request, admin_only=True)
    if err:
        return err
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    data   = _json.loads(request.body)
    name   = data.get('name', '').strip()
    photo  = data.get('photo', '')   # base64 data URL: "data:image/jpeg;base64,..."
    if not name or not photo:
        return JsonResponse({'error': 'name and photo required'}, status=400)
    photos = _load_photos()
    photos[name] = photo
    _save_photos(photos)
    return JsonResponse({'ok': True, 'name': name})


def attendance_face_photo_get(request, name):
    """Return the stored face photo for a person."""
    _, err = _api_auth(request)
    if err:
        return err
    photos = _load_photos()
    if name not in photos:
        return JsonResponse({'error': 'No photo'}, status=404)
    return JsonResponse({'ok': True, 'name': name, 'photo': photos[name]})


def attendance_face_photos_all(request):
    """Return all face photos as {name: dataURL}."""
    _, err = _api_auth(request)
    if err:
        return err
    return JsonResponse(_load_photos())


# ── HSE: SJN Portal ────────────────────────────────────────────────────────

@login_required
def hse_sjn_portal(request):
    """Serve the SJN O&M portal inside the HDEC shell."""
    user = get_user(request)
    return render(request, 'core/hse_sjn_portal.html', _ctx(request, {
        'page_title': 'SJN Portal',
        'category_color': '#22c55e',
    }))
